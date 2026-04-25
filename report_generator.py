"""
report_generator.py  ─  Excel帳票出力モジュール（完全版）
入荷記録 / 仕込み記録 / トレース / 月別集計 / 全件トレース
"""
from pathlib import Path
from datetime import date, datetime
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("data/reports")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── スタイル定数 ────────────────────────────────────────────────
C_HDR_BG  = "1565c0"
C_HDR_FG  = "ffffff"
C_TITLE   = "1a237e"
C_ROW_ALT = "e8f1fb"
C_TOTAL   = "fff9c4"
C_WARN    = "ffebee"
C_OK      = "e8f5e9"
C_BORDER  = "b0bec5"


def _side(): return Side(style="thin", color=C_BORDER)

def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())

def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="メイリオ", bold=True, color=C_HDR_FG, size=10)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, val, align="left", bg=None, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="メイリオ", size=9, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border    = _border()
    if bg:      c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt: c.number_format = num_fmt
    return c

def _title(ws, title, subtitle=""):
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 16
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name="メイリオ", bold=True, size=14, color=C_TITLE)
    if subtitle:
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = Font(name="メイリオ", size=9, color="607d8b")
    d = ws.cell(row=2, column=10, value=f"出力日: {date.today().strftime('%Y年%m月%d日')}")
    d.font = Font(name="メイリオ", size=9, color="607d8b")
    d.alignment = Alignment(horizontal="right")

def _parse_others(json_str: str) -> str:
    if not json_str:
        return ""
    try:
        items = json.loads(json_str)
        return " / ".join(
            f"{i.get('name','')}({i.get('lot','─')}): {i.get('kg',0):.3f}kg"
            for i in items if i.get("name")
        )
    except Exception:
        return json_str


# ═══════════════════════════════════════════════════════════════
# 入荷記録帳票
# ═══════════════════════════════════════════════════════════════
def generate_arrival_report(arrivals: list) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入荷記録"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    _title(ws, "📦 原料入荷記録帳票", f"件数: {len(arrivals)}")

    HDRS = [
        ("入荷No",12),("入荷日",12),("メーカー",14),("ロットNo",14),
        ("原料種別",18),("袋数",7),("総量(kg)",10),
        ("搬入温度",10),("外観",10),("臭い",10),("包装",10),
        ("色調",10),("異物",10),("水分",10),("賞味期限",10),
        ("異常内容",22),("担当者",10),("備考",18),
    ]
    HR = 4
    ws.row_dimensions[HR].height = 28
    for ci, (h, w) in enumerate(HDRS, start=2):
        _hdr(ws, HR, ci, h, w)

    def _ck(v):
        v = v or ""
        if "OK" in v or "なし" in v: return "✓ " + v
        if "NG" in v or "あり（要" in v: return "✗ " + v
        return v

    for ri, a in enumerate(reversed(arrivals), start=HR + 1):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 20
        row = [
            (a.get("arrival_no",""),    "center", None),
            (a.get("arrival_date",""),  "center", None),
            (a.get("maker",""),         "left",   None),
            (a.get("lot_no","") or "未設定", "center", None),
            (a.get("material_type",""),"left",   None),
            (float(a.get("bags",0)),    "right",  "#,##0"),
            (float(a.get("total_kg",0)),"right",  "#,##0.0"),
            (_ck(a.get("transport_temp","")), "center", None),
            (_ck(a.get("appearance","")),     "center", None),
            (_ck(a.get("odor","")),           "center", None),
            (_ck(a.get("packaging","")),      "center", None),
            (_ck(a.get("color_check","")),    "center", None),
            (_ck(a.get("contamination","")),  "center", None),
            (_ck(a.get("moisture","")),       "center", None),
            (_ck(a.get("expiry_check","")),   "center", None),
            (a.get("abnormal_detail","") or "─", "left", None),
            (a.get("inspector",""),   "center", None),
            (a.get("remarks","") or "─", "left", None),
        ]
        for ci, (val, align, nf) in enumerate(row, start=2):
            row_bg = C_WARN if ci == 9 and "NG" in str(val) else bg
            _cell(ws, ri, ci, val, align, bg=row_bg, num_fmt=nf)

    ws.freeze_panes = f"B{HR+1}"
    ws.auto_filter.ref = f"B{HR}:{get_column_letter(len(HDRS)+1)}{HR}"

    path = OUTPUT_DIR / f"入荷記録_{_ts()}.xlsx"
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════
# 仕込み記録帳票
# ═══════════════════════════════════════════════════════════════
def generate_brewing_report(brewing: list) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "仕込み記録"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    _title(ws, "🧪 仕込み記録帳票", f"件数: {len(brewing)}")

    HDRS = [
        ("No",6),("仕込日",12),("品名",20),("メーカー",12),("主ロット",14),
        ("仕込量(kg)",10),("精粉(kg)",9),("海藻粉(kg)",9),("海藻ロット",12),
        ("デンプン(kg)",10),("デンプンロット",12),("石灰(kg)",9),("石灰水(ℓ)",9),
        ("その他添加物",30),("備考",20),
    ]
    HR = 4
    ws.row_dimensions[HR].height = 28
    for ci, (h, w) in enumerate(HDRS, start=2):
        _hdr(ws, HR, ci, h, w)

    tot_brew = tot_mat = tot_sea = tot_lime = 0.0
    for ri, b in enumerate(reversed(brewing), start=HR + 1):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 22
        bv = float(b.get("brew_amount",0) or 0)
        mv = float(b.get("material_kg",0) or 0)
        sv = float(b.get("seaweed_kg",0) or 0)
        lv = float(b.get("lime_kg",0) or 0)
        tot_brew += bv; tot_mat += mv; tot_sea += sv; tot_lime += lv

        row = [
            (b.get("no",""),               "center", None),
            (b.get("brew_date",""),         "center", None),
            (b.get("product_name",""),      "left",   None),
            (b.get("maker",""),             "left",   None),
            (b.get("lot_no","") or "未設定","center", None),
            (bv,  "right", "#,##0.00"),
            (mv,  "right", "#,##0.00"),
            (sv,  "right", "#,##0.00"),
            (b.get("seaweed_lot","") or "─","center", None),
            (float(b.get("starch_kg",0) or 0), "right", "#,##0.00"),
            (b.get("starch_lot","") or "─","center", None),
            (lv,  "right", "#,##0.00"),
            (float(b.get("lime_water_l",0) or 0), "right", "#,##0.0"),
            (_parse_others(b.get("other_additives","")), "left", None),
            (b.get("notes","") or "─",     "left",   None),
        ]
        for ci, (val, align, nf) in enumerate(row, start=2):
            _cell(ws, ri, ci, val, align, bg=bg, num_fmt=nf)

    # 合計行
    tr = HR + len(brewing) + 1
    ws.row_dimensions[tr].height = 22
    _cell(ws, tr, 2, "合計", "center", bg=C_TOTAL, bold=True)
    for ci, val in [(7, tot_brew),(8, tot_mat),(9, tot_sea),(13, tot_lime)]:
        c = _cell(ws, tr, ci, val, "right", bg=C_TOTAL, num_fmt="#,##0.00", bold=True)

    ws.freeze_panes = f"B{HR+1}"
    ws.auto_filter.ref = f"B{HR}:{get_column_letter(len(HDRS)+1)}{HR}"

    path = OUTPUT_DIR / f"仕込み記録_{_ts()}.xlsx"
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════
# トレース帳票（フォワード / バックワード）
# ═══════════════════════════════════════════════════════════════
def generate_trace_report(results: list, trace_type: str, keyword: str) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "トレース結果"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    _title(ws, f"🔍 原料トレース帳票（{trace_type}）",
           f"検索条件: {keyword} ／ 件数: {len(results)}")

    if not results:
        ws.cell(row=4, column=2, value="該当データなし")
        path = OUTPUT_DIR / f"トレース_{_ts()}.xlsx"
        wb.save(path); return path

    cols = list(results[0].keys())
    WIDTHS = {"仕込No":7,"仕込日":12,"品名":20,"仕込量(kg)":10,"役割":14,
              "使用ロット":14,"使用量(kg)":10,"入荷No":12,"メーカー":14,
              "入荷日":12,"外観検査":10,"担当者":10}
    HR = 4
    ws.row_dimensions[HR].height = 28
    for ci, col in enumerate(cols, start=2):
        _hdr(ws, HR, ci, col, WIDTHS.get(col, 12))

    for ri, row in enumerate(results, start=HR + 1):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 20
        for ci, col in enumerate(cols, start=2):
            val = row.get(col, "")
            align = "right" if isinstance(val, (int, float)) else "center"
            nf = "#,##0.00" if isinstance(val, float) else None
            _cell(ws, ri, ci, val, align, bg=bg, num_fmt=nf)

    path = OUTPUT_DIR / f"トレース_{trace_type}_{_ts()}.xlsx"
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════
# 全件トレースレポート（HACCP一括出力）
# ═══════════════════════════════════════════════════════════════
def generate_full_trace_report(arrivals: list, brewing: list) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全件トレース"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    _title(ws, "📋 全件原料トレースレポート（HACCP用）",
           f"仕込み件数: {len(brewing)}件 ／ 入荷件数: {len(arrivals)}件")

    lot_map = {str(a.get("lot_no","")).strip(): a for a in arrivals if a.get("lot_no")}
    ano_map  = {str(a.get("arrival_no","")).strip(): a for a in arrivals}

    HDRS = [
        ("仕込No",7),("仕込日",12),("品名",20),("仕込量(kg)",10),
        ("役割",16),("使用ロット",14),("使用量(kg)",10),
        ("入荷No",12),("メーカー",14),("入荷日",12),("外観",10),("担当者",10),
    ]
    HR = 4
    ws.row_dimensions[HR].height = 28
    for ci, (h, w) in enumerate(HDRS, start=2):
        _hdr(ws, HR, ci, h, w)

    ri = HR + 1
    for b in brewing:
        # 使用原料リスト
        used = []
        for lot_k, role, kg_k in [
            ("lot_no","主原料（精粉）","material_kg"),
            ("seaweed_lot","海藻粉","seaweed_kg"),
            ("starch_lot","加工デンプン","starch_kg"),
        ]:
            lot_v = str(b.get(lot_k,"")).strip()
            kg_v  = float(b.get(kg_k, 0) or 0)
            used.append({"role": role, "lot": lot_v, "kg": kg_v})

        oa = b.get("other_additives","")
        if oa:
            try:
                for o in json.loads(oa):
                    used.append({"role": o.get("name","添加物"),
                                 "lot": str(o.get("lot","")).strip(),
                                 "kg": float(o.get("kg",0) or 0)})
            except Exception:
                pass

        for u in used:
            lot_v = u["lot"]
            arr_m = lot_map.get(lot_v) or ano_map.get(lot_v)

            bg = C_ROW_ALT if ri % 2 == 0 else None
            ws.row_dimensions[ri].height = 18
            row_data = [
                (b.get("no",""),             "center"),
                (b.get("brew_date",""),       "center"),
                (b.get("product_name",""),    "left"),
                (float(b.get("brew_amount",0) or 0), "right"),
                (u["role"],                   "left"),
                (lot_v or "─",               "center"),
                (u["kg"],                     "right"),
                (arr_m.get("arrival_no","─") if arr_m else "不明", "center"),
                (arr_m.get("maker","不明")   if arr_m else "不明", "left"),
                (arr_m.get("arrival_date","─") if arr_m else "─", "center"),
                (arr_m.get("appearance","─") if arr_m else "─",   "center"),
                (arr_m.get("inspector","─") if arr_m else "─",    "center"),
            ]
            for ci, (val, align) in enumerate(row_data, start=2):
                nf = "#,##0.00" if isinstance(val, float) else None
                _cell(ws, ri, ci, val, align, bg=bg, num_fmt=nf)
            ri += 1

    ws.freeze_panes = f"B{HR+1}"
    ws.auto_filter.ref = f"B{HR}:{get_column_letter(len(HDRS)+1)}{HR}"

    path = OUTPUT_DIR / f"全件トレース_{_ts()}.xlsx"
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════
# 月別集計帳票
# ═══════════════════════════════════════════════════════════════
def generate_monthly_report(monthly: list, brewing_all: list) -> Path:
    wb = openpyxl.Workbook()

    # シート1: 集計
    ws = wb.active
    ws.title = "月別集計"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    _title(ws, "📊 生産集計帳票")

    HDRS = [("期間",14),("件数",8),("仕込量(kg)",12),("精粉(kg)",11),
            ("海藻粉(kg)",11),("デンプン(kg)",11),("歩留まり(倍)",12)]
    HR = 4
    ws.row_dimensions[HR].height = 28
    for ci, (h, w) in enumerate(HDRS, start=2): _hdr(ws, HR, ci, h, w)

    for ri, row in enumerate(monthly, start=HR + 1):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 20
        vals = list(row.values())
        for ci, val in enumerate(vals, start=2):
            align = "center" if ci == 2 else "right"
            nf = "#,##0.00" if isinstance(val, float) else None
            _cell(ws, ri, ci, val, align, bg=bg, num_fmt=nf)

    # シート2: 仕込み明細
    ws2 = wb.create_sheet("仕込み明細")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 2
    _title(ws2, "仕込み記録 全明細")

    HDRS2 = [("No",6),("仕込日",12),("品名",20),("メーカー",12),
              ("主ロット",12),("仕込量(kg)",10),("精粉(kg)",9),("備考",20)]
    for ci, (h, w) in enumerate(HDRS2, start=2): _hdr(ws2, 4, ci, h, w)
    for ri, b in enumerate(reversed(brewing_all), start=5):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws2.row_dimensions[ri].height = 18
        vals = [b.get("no",""), b.get("brew_date",""), b.get("product_name",""),
                b.get("maker",""), b.get("lot_no","") or "未設定",
                float(b.get("brew_amount",0) or 0),
                float(b.get("material_kg",0) or 0), b.get("notes","")]
        for ci, val in enumerate(vals, start=2):
            align = "right" if isinstance(val, float) else "center" if ci <= 6 else "left"
            _cell(ws2, ri, ci, val, align, bg=bg,
                  num_fmt="#,##0.00" if isinstance(val, float) else None)

    path = OUTPUT_DIR / f"生産集計_{_ts()}.xlsx"
    wb.save(path)
    return path


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")
