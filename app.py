# app.py
import streamlit as st
import pandas as pd
import json
import time
import base64
import re
from io import BytesIO
from datetime import datetime, date, timedelta
import traceback
import plotly.graph_objects as go
import plotly.express as px

try:
    import sheets
except ImportError:
    pass

# Excel出力用
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 資材画像アップロード用
try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

st.set_page_config(page_title="食品工場 製造ERP", page_icon="🏭", layout="wide", initial_sidebar_state="expanded")

# ════════════════════════════════════════════════════════════════
#  UI/UX CSS & 共通ヘルパー
# ════════════════════════════════════════════════════════════════
st.markdown("""
<style>
:root {
    --c-bg: #f4f6f7; --c-surface: #ffffff; --c-primary: #0f766e; --c-primary-hover: #0b5c56;
    --c-primary-soft: #e6f2f1; --c-secondary: #1e293b; --c-muted: #64748b; --c-border: #dbe2e6;
    --c-input-border: #a8b3ba; --c-danger: #b91c1c; --c-danger-bg: #fdf1f1; --c-success: #15803d;
    --c-water: #94a3b8; --radius-lg: 16px; --radius-md: 10px; --radius-sm: 8px;
    --shadow-card: 0 2px 6px -1px rgba(15,23,42,0.08);
}
html, body, .stApp { background-color: var(--c-bg) !important; font-family: -apple-system, sans-serif !important; }
h1, h2, h3, h4, p, span, div, label { color: var(--c-secondary); }
.block-container { padding-top: 1.5rem !important; max-width: 1280px; }
.main-header { background: var(--c-surface); padding: 18px 24px; border-radius: var(--radius-lg); margin-bottom: 24px; box-shadow: var(--shadow-card); border-left: 8px solid var(--c-primary); }
.main-header h1 { font-size: 1.6rem !important; margin: 0 0 6px 0 !important; font-weight: 900 !important; }
.form-card { background: var(--c-surface); border-radius: var(--radius-lg); padding: 24px; margin-bottom: 24px; box-shadow: var(--shadow-card); border: 1px solid #e2e8f0; }
.section-title { font-size: 1.25rem; font-weight: 900; margin-bottom: 20px; border-bottom: 3px solid var(--c-border); padding-bottom: 8px; }
div[data-testid="stRadio"] > div { display: flex; flex-wrap: wrap; gap: 8px !important; }
div[data-testid="stRadio"] label { background-color: #ffffff; padding: 10px 16px !important; border-radius: var(--radius-sm); border: 2px solid var(--c-border) !important; cursor: pointer; text-align: center; flex: 1 1 auto; justify-content: center; min-width: 80px; transition: all 0.15s ease; }
div[data-testid="stRadio"] label p { font-size: 1.0rem !important; font-weight: 700 !important; color: var(--c-secondary) !important; }
div[data-testid="stRadio"] label:has(input:checked) { background-color: var(--c-primary) !important; border-color: var(--c-primary-hover) !important; box-shadow: 0 3px 10px rgba(15,118,110,0.25) !important; transform: translateY(-1px); }
div[data-testid="stRadio"] label:has(input:checked) * { color: #ffffff !important; font-weight: 900 !important; fill: #ffffff !important; }
div[data-baseweb="input"] { background-color: #ffffff !important; border: 3px solid var(--c-input-border) !important; border-radius: var(--radius-md) !important; }
div[data-baseweb="input"]:focus-within { border-color: var(--c-primary) !important; box-shadow: 0 0 0 5px rgba(15,118,110,0.18) !important; }
div[data-testid="stNumberInputContainer"] { min-height: 50px !important; background-color: #f8fafc !important; }
div[data-testid="stNumberInputContainer"] input { font-size: 1.2rem !important; font-weight: 800 !important; text-align: center !important; }
.stButton button { border-radius: var(--radius-sm) !important; font-weight: 800 !important; padding: 14px 20px !important; min-height: 52px !important; border: 2px solid var(--c-input-border) !important; }
.stButton button[kind="primary"] { background: var(--c-primary) !important; color: #ffffff !important; border: none !important; box-shadow: 0 4px 12px rgba(15,118,110,0.3) !important; }
.stButton button[kind="primary"]:hover { background: var(--c-primary-hover) !important; transform: translateY(-2px); }
.ratio-btn-container .stButton button { min-height: 38px !important; padding: 4px 6px !important; font-size: 0.95rem !important; background: #f8fafc !important; border: 1px solid #cbd5e1 !important; border-radius: 6px !important; }
.ratio-btn-container .stButton button:hover { background: var(--c-primary-soft) !important; border-color: var(--c-primary) !important; color: var(--c-primary) !important; }
.ratio-btn-container .stButton button[kind="primary"] { background: var(--c-primary) !important; border: none !important; color: #ffffff !important; }
[data-testid="stSidebar"] { background-color: #f8fafc !important; border-right: 2px solid var(--c-border); padding-top: 1rem; }
[data-testid="stSidebar"] div[role="radiogroup"] label { background: #ffffff !important; border: 2px solid var(--c-border) !important; padding: 14px 16px !important; border-radius: var(--radius-md) !important; margin-bottom: 0 !important; }
</style>
""", unsafe_allow_html=True)

def card_start(): st.markdown('<div class="form-card">', unsafe_allow_html=True)
def card_end(): st.markdown('</div>', unsafe_allow_html=True)
def sec_title(title): st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
#  データロード & パーサー
# ════════════════════════════════════════════════════════════════
def refresh(): st.cache_data.clear(); st.rerun()

@st.cache_data(ttl=60)
def load_all_datasets():
    import sheets
    return {
        "arrivals": sheets.load_arrivals(), "brewing": sheets.load_brewing(), "adjustments": sheets.load_adjustments(),
        "supplies": sheets.load_supplies(), "supply_logs": sheets.load_supply_logs(),
        "materials": sheets.load_materials(), "makers": sheets.load_makers(), "inspectors": sheets.load_inspectors(),
        "order_points": sheets.load_order_points(), "recipes": sheets.load_recipes(), "recipe_logs": sheets.load_recipe_logs(),
        "grades": sheets.load_grades() if hasattr(sheets, "load_grades") else None,
        "purchase_orders": sheets.load_purchase_orders() if hasattr(sheets, "load_purchase_orders") else None
    }

try:
    import sheets
    dataset = load_all_datasets()
    arrivals, brewing, adjustments = dataset.get("arrivals", []), dataset.get("brewing", []), dataset.get("adjustments", [])
    supplies, supply_logs = dataset.get("supplies", []), dataset.get("supply_logs", [])
    materials, makers, inspectors = dataset.get("materials", []), dataset.get("makers", []), dataset.get("inspectors", [])
    order_points, recipes_raw, recipe_logs = dataset.get("order_points", {}), dataset.get("recipes", []), dataset.get("recipe_logs", [])
    grades_data, purchase_orders_data = dataset.get("grades"), dataset.get("purchase_orders")
except Exception:
    st.error("🚨 データの読み込みに失敗しました。")
    st.stop()

def parse_op_data(raw_val):
    pt, wt = 0, 20
    try:
        if isinstance(raw_val, str):
            if raw_val.strip().startswith("{"):
                d = json.loads(raw_val)
                pt, wt = int(float(d.get("pt", 0))), int(float(d.get("wt", 20)))
            else:
                m_pt = re.search(r"発注点:([\d\.]+)袋", raw_val); m_wt = re.search(r"重量:([\d\.]+)kg", raw_val)
                if m_pt: pt = int(float(m_pt.group(1)))
                if m_wt: wt = int(float(m_wt.group(1)))
        else: pt = int(float(raw_val))
    except: pass
    return pt, wt

def get_material_image(op_dict, mat_name): return op_dict.get(f"__IMAGE_{mat_name}__", "")
def get_active_lots_from_master(op_dict, mat_name):
    v = op_dict.get(f"__ACTIVE_LOTS_{mat_name}__", "")
    if v:
        try: return json.loads(v)
        except: pass
    return []

def parse_lime_config(op_dict, product_name=None):
    c = {"start_month": 6, "end_month": 9, "add_ratio": 0.01, "reason": "夏場の高温対策"}
    try:
        v = op_dict.get(f"__LIME_CONFIG_{product_name}__", "") if product_name else ""
        if not v: v = op_dict.get("__LIME_CONFIG__", "")
        if v:
            m_s = re.search(r"開始:(\d+)月", v); m_e = re.search(r"終了:(\d+)月", v); m_r = re.search(r"割合:([\d\.]+)", v); m_reason = re.search(r"理由:(.+?)(?:,|$)", v)
            if m_s: c["start_month"] = int(m_s.group(1))
            if m_e: c["end_month"] = int(m_e.group(1))
            if m_r: c["add_ratio"] = float(m_r.group(1))
            if m_reason: c["reason"] = m_reason.group(1).strip()
    except: pass
    return c

def parse_grade_list(op_dict):
    if grades_data is not None: return grades_data
    try:
        v = op_dict.get("__GRADE_LIST__", "")
        if v: return [str(x).strip() for x in json.loads(v)] if v.strip().startswith("[") else [str(x).strip() for x in v.split(",") if str(x).strip()]
    except: pass
    return []

def safe_parse_recipe(r_val):
    if not r_val: return []
    if isinstance(r_val, str) and r_val.strip().startswith("["):
        try: return json.loads(r_val)
        except: pass
    items = []
    if isinstance(r_val, str):
        for p in r_val.split(","):
            if ":" in p:
                mat, rat = p.split(":", 1)
                try: items.append({"原料名": mat.strip(), "比率": float(rat.replace("%", "").strip())})
                except: pass
    return items

def safe_parse_seasoning_recipe(r_val):
    if not r_val: return []
    if isinstance(r_val, str) and r_val.strip().startswith("["):
        try: return json.loads(r_val)
        except: pass
    items = []
    if isinstance(r_val, str):
        for p in r_val.split(","):
            if ":" in p:
                mat, rat = p.split(":", 1)
                try: items.append({"原料名": mat.strip(), "希釈倍率": float(rat.replace("倍", "").strip())})
                except: pass
    return items

def parse_brewing_ingredients(r_val):
    if not r_val: return []
    if isinstance(r_val, str) and r_val.strip().startswith("["):
        try: return json.loads(r_val)
        except: pass
    items = []
    if isinstance(r_val, str):
        for p in r_val.split(","):
            m = re.match(r"(.+?):([\d\.]+)kg\((.+?)\)", p.strip())
            if m: items.append({"原料名": m.group(1).strip(), "kg": float(m.group(2)), "lot": m.group(3).strip()})
    return items

def is_lime_boost_active(cfg, t_date=None):
    if t_date is None: t_date = date.today()
    m, s, e = t_date.month, int(cfg.get("start_month", 6)), int(cfg.get("end_month", 9))
    return s <= m <= e if s <= e else (m >= s or m <= e)

def save_grade_list(op_dict, g_list):
    if hasattr(sheets, "save_grades"): sheets.save_grades(g_list)
    else: d = dict(op_dict); d["__GRADE_LIST__"] = ", ".join(g_list); sheets.save_order_points(d)

def is_konjac_material(name):
    s = str(name); s_hira = "".join(chr(ord(c)-0x60) if "ァ"<=c<="ヶ" else c for c in s)
    return ("こんにゃく" in s_hira) or ("蒟蒻" in s) or ("konnyaku" in s.lower())

def parse_purchase_orders(op_dict):
    if purchase_orders_data is not None: return purchase_orders_data
    try:
        v = op_dict.get("__PURCHASE_ORDERS__", "")
        if v and v.startswith("["): return json.loads(v)
    except: pass
    return []

def save_purchase_orders(op_dict, o_list):
    if hasattr(sheets, "save_purchase_orders"): sheets.save_purchase_orders(o_list)
    else: d = dict(op_dict); d["__PURCHASE_ORDERS__"] = json.dumps(o_list, ensure_ascii=False); sheets.save_order_points(d)

BIG_CAT_ICONS = {"プラント": "🏭", "OKM": "🟦", "手詰め": "✋"}
SUB_CAT_ICONS = {"白": "⚪", "黒": "⚫", "耐冷": "❄️", "ショクカイ": "🍽️", "めん": "🍜", "おでん": "🍢", "その他": "📦"}
_ICON_POOL = ["🔵", "🟢", "🟡", "🟣", "🟠", "🔴", "🟤", "🔷", "🔶", "🔹", "🔸", "⬛", "⬜", "🟥", "🟩", "🟦"]
_PRODUCT_ICON_POOL = ["🍥", "🥢", "🌿", "🎍", "🧊", "🍡", "🧵", "🏷️", "📌", "🧺", "🔖", "🧫"]

def _deterministic_icon(name, pool): return pool[sum(ord(ch) for ch in str(name)) % len(pool)]
def big_cat_icon(name): return BIG_CAT_ICONS.get(name, _deterministic_icon(name, _ICON_POOL))
def sub_cat_icon(name): return SUB_CAT_ICONS.get(name, _deterministic_icon(name, _ICON_POOL))
def product_icon(name): return _deterministic_icon(name, _PRODUCT_ICON_POOL)

def fmt_kg(val):
    if val is None or val == "": return "0"
    try: v = float(val); return f"{int(v)}" if v.is_integer() else f"{v:.2f}".rstrip('0').rstrip('.')
    except: return str(val)

def fmt_df_numeric(df, cols):
    d = df.copy()
    for c in cols:
        if c in d.columns: d[c] = d[c].apply(fmt_kg)
    return d

# ════════════════════════════════════════════════════════════════
#  在庫計算 (誤差丸めによる完全0化)
# ════════════════════════════════════════════════════════════════
def get_inventory():
    inv = {}
    for a in arrivals:
        ano = str(a.get("入荷No", "")).strip()
        if not ano: continue
        inv[ano] = {
            "入荷No": ano, "入荷日": str(a.get("入荷日", "")).strip() or "-", "ロットNo": str(a.get("ロットNo", "")).strip(), "原料種別": str(a.get("原料種別", "")).strip(), 
            "メーカー": str(a.get("メーカー", "")).strip(), "グレード": str(a.get("グレード", "")).strip(),
            "1袋重量": round(float(a.get("1袋重量(kg)") or 20.0), 3), "入荷袋数": round(float(a.get("袋数") or 0.0), 3), "使用量(kg)": 0.0, "調整袋数": 0.0
        }
    for b in brewing:
        oa = b.get("その他添加物", "")
        if oa:
            items = parse_brewing_ingredients(oa)
            for item in items:
                t_lot, t_kg = str(item.get("lot", "")).strip(), float(item.get("kg", 0.0))
                v_lots = [l for l in [re.sub(r'\(\d+%\)', '', x).strip() for x in t_lot.split(",")] if l and l != "─"]
                if v_lots:
                    kl = t_kg / len(v_lots)
                    for l in v_lots:
                        for v in inv.values():
                            if v["ロットNo"] == l: v["使用量(kg)"] += kl
    for adj in adjustments:
        ano = str(adj.get("入荷No", "")).strip()
        if ano in inv: inv[ano]["調整袋数"] += float(adj.get("調整袋数") or 0.0)
    
    for v in inv.values():
        bpk = v["1袋重量"] if v["1袋重量"] > 0 else 20.0
        v["使用袋数"] = round(v["使用量(kg)"] / bpk, 4)
        raw_bags = v["入荷袋数"] - v["使用袋数"] + v["調整袋数"]
        if abs(raw_bags) < 0.01: raw_bags = 0.0
        v["現在庫(袋)"] = max(round(raw_bags, 3), 0.0)
        v["現在庫(kg)"] = round(v["現在庫(袋)"] * bpk, 3)
    return inv

inventory_data = get_inventory()
type_totals_kg, type_totals_bag = {}, {}
for v in inventory_data.values():
    m = v["原料種別"]
    type_totals_kg[m] = type_totals_kg.get(m, 0.0) + v["現在庫(kg)"]
    type_totals_bag[m] = type_totals_bag.get(m, 0.0) + v["現在庫(袋)"]

def _get_active_lots(mat):
    """在庫が0.01袋以上あるロットのみを返す"""
    o = []
    for v in inventory_data.values():
        if v["原料種別"] == mat and v["現在庫(袋)"] > 0.01 and v["ロットNo"] not in o: o.append(v["ロットNo"])
    return o

def get_lots_for_material(mat):
    l = [v for v in inventory_data.values() if v["原料種別"] == mat]
    l.sort(key=lambda v: v["現在庫(袋)"], reverse=True)
    return l

def get_supply_inventory():
    inv = {s.get("資材ID"): float(s.get("初期在庫") or 0.0) for s in supplies}
    for log in supply_logs:
        sid, qty, act = log.get("資材ID"), float(log.get("数量") or 0.0), log.get("処理")
        if sid in inv:
            if act == "入荷": inv[sid] += qty
            elif act == "使用": inv[sid] -= qty
    return inv

# ════════════════════════════════════════════════════════════════
#  カスタムUIコンポーネント (タップで閉じる＆ロット整理対応)
# ════════════════════════════════════════════════════════════════
def render_amount_adjuster(title, calc_val, p_key):
    st.markdown(f"<div style='font-size:1.05rem; font-weight:800; color:#475569; margin-bottom:4px;'>{title}</div>", unsafe_allow_html=True)
    lst_key = f"last_calc_{p_key}"
    last_calc, calc_val = st.session_state.get(lst_key, None), round(calc_val, 2)
    if (last_calc is None) or (abs(float(last_calc) - float(calc_val)) > 1e-6):
        st.session_state[p_key] = st.session_state[lst_key] = calc_val
    if p_key not in st.session_state: st.session_state[p_key] = calc_val

    st.markdown(f"""
    <div style="background-color:#f0f9ff; border:2px solid #38bdf8; border-radius:8px; padding:10px; margin-bottom:8px; text-align:center; box-shadow:inset 0 1px 3px rgba(0,0,0,0.06);">
        <span style="font-size:2.0rem; font-weight:900; color:#0284c7;">{fmt_kg(st.session_state[p_key])}</span>
        <span style="font-size:1.0rem; color:#0369a1; font-weight:700; margin-left:4px;">kg</span>
    </div>
    """, unsafe_allow_html=True)
    return st.number_input("微調整", min_value=0.0, step=0.1, key=p_key, label_visibility="collapsed")


def render_lot_selector(mat_name, lot_key):
    """
    在庫0のロットは一切出さず、マスタで指定されたロットを最優先で表示。
    指定されていない在庫ありロットは「その他の在庫ありロット」に格納し、タップで確定＆閉じる。
    """
    master_lots = get_active_lots_from_master(order_points, mat_name)
    all_active_lots = _get_active_lots(mat_name)
    
    # マスタで指定されているが在庫がないロットは除外する（安全性のため）
    valid_master_lots = [l for l in master_lots if l in all_active_lots]
    # マスタで指定されていない在庫ありロット
    other_lots = [l for l in all_active_lots if l not in valid_master_lots]
    
    # 初期値の設定
    curr_val = st.session_state.get(lot_key, valid_master_lots[0] if valid_master_lots else (other_lots[0] if other_lots else "─"))
    pop_label = f"✅ 選択済: {curr_val}" if curr_val not in ["─", ""] else "⚠️ ロット未選択 (タップ)"
    
    st.markdown(f"<div style='font-size:1.0rem; font-weight:800; color:#475569; margin-bottom:6px;'>📦 ロット選択</div>", unsafe_allow_html=True)
    with st.popover(pop_label, use_container_width=True):
        st.markdown(f"**📦 {mat_name} のロット選択**")
        d_map = {v["ロットNo"]: v["入荷日"] for v in inventory_data.values() if v["原料種別"] == mat_name}
        
        if valid_master_lots:
            st.caption("📌 マスタで指定された使用中ロット")
            for opt in valid_master_lots:
                if st.button(f"{opt} (入荷:{d_map.get(opt, '不明')})", key=f"btn_{lot_key}_{opt}", use_container_width=True):
                    st.session_state[lot_key] = opt
                    st.rerun()
                    
        if other_lots:
            if valid_master_lots:
                with st.expander("📦 その他の在庫ありロット"):
                    for opt in other_lots:
                        if st.button(f"{opt} (入荷:{d_map.get(opt, '不明')})", key=f"btn_{lot_key}_{opt}", use_container_width=True):
                            st.session_state[lot_key] = opt
                            st.rerun()
            else:
                st.caption("📦 在庫ありロット")
                for opt in other_lots:
                    if st.button(f"{opt} (入荷:{d_map.get(opt, '不明')})", key=f"btn_{lot_key}_{opt}", use_container_width=True):
                        st.session_state[lot_key] = opt
                        st.rerun()
                        
        if not valid_master_lots and not other_lots:
            st.caption("選択可能なロットがありません。")
            
        st.divider()
        m_in = st.text_input("✏️ ロット手入力", key=f"txt_{lot_key}")
        if st.button("手入力で確定", key=f"btn_manual_{lot_key}", use_container_width=True):
            if m_in.strip():
                st.session_state[lot_key] = m_in.strip()
                st.rerun()
                
    return st.session_state.get(lot_key, "─")


def render_operator_selector(operator_key):
    if operator_key not in st.session_state: st.session_state[operator_key] = inspectors[0] if inspectors else "未登録"
    with st.popover(f"👨‍🏭 担当者: {st.session_state[operator_key]}", use_container_width=True):
        for insp in inspectors:
            if st.button(insp, key=f"btn_insp_{operator_key}_{insp}", use_container_width=True):
                st.session_state[operator_key] = insp
                st.rerun()
    return st.session_state[operator_key]


def render_excel_history_editor(full_records, filtered_df, id_col, editable_cols, numeric_cols, save_func, key_prefix):
    if filtered_df.empty:
        st.info("データがありません。")
        return
    edit_df = filtered_df[[id_col] + [c for c in editable_cols if c != id_col]].copy().reset_index(drop=True)
    for c in numeric_cols:
        if c in edit_df.columns: edit_df[c] = pd.to_numeric(edit_df[c], errors="coerce").fillna(0.0)
    edit_df[id_col] = edit_df[id_col].astype(str)
    
    st.caption("💡 セルをタップして直接編集できます。行左端のチェックで選択し🗑️で削除できます。")
    col_cfg = {id_col: st.column_config.TextColumn(id_col, disabled=True)}
    for c in numeric_cols:
        if c in edit_df.columns: col_cfg[c] = st.column_config.NumberColumn(c, format="%.2f")

    edited_df = st.data_editor(edit_df, num_rows="dynamic", use_container_width=True, hide_index=True, key=f"{key_prefix}_editor", column_config=col_cfg)
    diff_key = f"{key_prefix}_diff_pending"

    c_check, c_cancel = st.columns([2, 1])
    if c_check.button("🔍 変更内容を確認する", key=f"{key_prefix}_check_btn", use_container_width=True):
        orig_ids = set(edit_df[id_col])
        edited_clean = edited_df.copy()
        edited_clean[id_col] = edited_clean[id_col].astype(str).str.strip()
        valid_edited = edited_clean[edited_clean[id_col] != ""]
        changed_rows = [row[id_col] for _, row in valid_edited.iterrows() if row[id_col] in orig_ids and any(str(row[c]) != str(edit_df[edit_df[id_col] == row[id_col]].iloc[0][c]) for c in editable_cols if c != id_col)]
        
        st.session_state[diff_key] = {
            "changed": changed_rows, "deleted": sorted(orig_ids - set(valid_edited[id_col])),
            "blank_new_rows": len(edited_clean) - len(valid_edited), "edited_df": edited_clean
        }
        st.rerun()

    if c_cancel.button("❌ 取消", key=f"{key_prefix}_cancel_btn", use_container_width=True):
        st.session_state.pop(diff_key, None)
        st.rerun()

    diff = st.session_state.get(diff_key)
    if diff:
        n_c, n_d, n_b = len(diff["changed"]), len(diff["deleted"]), diff["blank_new_rows"]
        if n_c == 0 and n_d == 0 and n_b == 0: st.info("変更はありませんでした。")
        else:
            msg = []
            if n_c: msg.append(f"✏️ 更新 {n_c}件")
            if n_d: msg.append(f"🗑️ 削除 {n_d}件")
            st.markdown(f'<div style="background:{"#fdf1f1" if n_d else "#e6f2f1"}; border:2px solid {"#b91c1c" if n_d else "#0f766e"}; border-radius:10px; padding:14px; margin:10px 0;">{"<br>".join(msg)}</div>', unsafe_allow_html=True)

            if (n_c or n_d) and st.button("✅ この内容で確定保存する", type="primary", key=f"{key_prefix}_confirm_btn", use_container_width=True):
                id_to_record = {str(r.get(id_col)): dict(r) for r in full_records}
                valid_edited = diff["edited_df"][diff["edited_df"][id_col] != ""]
                for _, row in valid_edited.iterrows():
                    if row[id_col] in id_to_record:
                        for c in editable_cols:
                            if c != id_col: id_to_record[row[id_col]][c] = float(row[c]) if c in numeric_cols else row[c]
                for rid in diff["deleted"]: id_to_record.pop(rid, None)
                save_func(list(id_to_record.values()))
                st.session_state.pop(diff_key, None)
                st.success("保存しました。"); time.sleep(1.5); refresh()


def render_lot_inventory_manager(active_inv):
    st.caption("💡 棚卸し・在庫調整はこの表で行います。「現在庫(袋)」の数値を実際の数に書き換えてください。**「0」にすれば確実にゼロになります。**")
    rows = sorted(active_inv, key=lambda v: v["入荷日"])
    orig_map = {v["入荷No"]: v for v in rows}

    df_edit = pd.DataFrame([{"入荷No": v["入荷No"], "入荷日": v["入荷日"], "原料種別": v["原料種別"], "メーカー": v["メーカー"], "ロットNo": v["ロットNo"], "現在庫(袋)": round(v["現在庫(袋)"], 3), "🗑️ 削除(在庫0に)": False} for v in rows])
    col_cfg = {c: st.column_config.TextColumn(c, disabled=True) for c in ["入荷No", "入荷日", "原料種別", "メーカー", "ロットNo"]}
    col_cfg["現在庫(袋)"] = st.column_config.NumberColumn("現在庫(袋)【編集可】", format="%.3f", min_value=0.0, step=0.1)

    edited_df = st.data_editor(df_edit, use_container_width=True, hide_index=True, num_rows="fixed", key="lot_inv_editor", column_config=col_cfg)
    diff_key = "lot_inv_diff_pending"
    
    c_check, c_cancel = st.columns([2, 1])
    if c_check.button("🔍 変更内容を確認する", key="lot_inv_check_btn", use_container_width=True):
        changes = []
        for _, r in edited_df.iterrows():
            ano = str(r["入荷No"])
            orig = orig_map.get(ano)
            if not orig: continue
            del_flag = bool(r.get("🗑️ 削除(在庫0に)", False))
            theo = orig["現在庫(袋)"]
            new_bags = 0.0 if del_flag else max(0.0, float(r["現在庫(袋)"]))
            diff = round(new_bags - theo, 6)
            if del_flag or abs(diff) > 0.005:
                changes.append({"入荷No": ano, "ロットNo": orig["ロットNo"], "原料種別": orig["原料種別"], "旧在庫": theo, "新在庫": new_bags, "差分": diff, "削除": del_flag})
        st.session_state[diff_key] = changes
        st.rerun()

    if c_cancel.button("❌ 取消", key="lot_inv_cancel_btn", use_container_width=True):
        st.session_state.pop(diff_key, None); st.rerun()

    changes = st.session_state.get(diff_key)
    if changes is not None:
        if not changes: st.info("変更はありませんでした。")
        else:
            msg_lines = [f"{'🗑️ 0設定' if c['削除'] or c['新在庫']==0 else '✏️ 編集'} {c['原料種別']} ロット:{c['ロットNo']} {fmt_kg(c['旧在庫'])}袋 → {fmt_kg(c['新在庫'])}袋" for c in changes]
            st.markdown(f'<div style="background:#fdf1f1; border:2px solid #b91c1c; border-radius:10px; padding:14px; margin:10px 0;">{"<br>".join(msg_lines)}</div>', unsafe_allow_html=True)
            reason_txt = st.text_input("変更理由", key="lot_inv_reason")
            op = render_operator_selector("lot_inv_op")
            if st.button("✅ 確定保存する（履歴に記録）", type="primary", key="lot_inv_confirm_btn", use_container_width=True):
                if hasattr(sheets, "append_adjustment"):
                    for c in changes:
                        sheets.append_adjustment({
                            "調整ID": f"ADJ-{datetime.now().strftime('%Y%m%d%H%M%S%f')}", "入荷No": c["入荷No"], "調整日": str(date.today()),
                            "調整袋数": c["差分"], "理由": f"【ロット現在庫編集】{fmt_kg(c['旧在庫'])}→{fmt_kg(c['新在庫'])} {reason_txt}",
                            "担当者": op, "登録日時": datetime.now().isoformat()
                        })
                    st.success(f"{len(changes)}件の変更を保存しました。")
                    st.session_state.pop(diff_key, None)
                    time.sleep(1.5); refresh()

# ════════════════════════════════════════════════════════════════
#  サイドバー
# ════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div style="font-size:1.6rem; font-weight:900; margin-bottom:1.5rem; color:#0f172a;">🏭 製造ERP</div>', unsafe_allow_html=True)
    page = st.radio("メニュー", [
        "🏭 製造仕込み", "📊 ダッシュボード", "📝 発注管理", "📥 入荷登録", "📦 在庫・棚卸", 
        "🧹 資材管理", "🔍 トレース", "📋 履歴・帳票", "📈 分析", "⚙️ マスタ設定"
    ], label_visibility="collapsed")
    st.markdown("---")
    if st.button("🔄 最新データに更新", use_container_width=True): refresh()


# ═══════════════════════════════════════════════════════════════
#  🏭 製造仕込み
# ═══════════════════════════════════════════════════════════════
if page == "🏭 製造仕込み":
    st.markdown('<div class="main-header"><h1>🏭 製造仕込み記録</h1><p>投入量は完全自動計算されます。微調整も可能です。</p></div>', unsafe_allow_html=True)
    card_start()
    brew_date = st.date_input("📅 仕込日", value=date.today())
    st.markdown("<br>", unsafe_allow_html=True)

    p_recipes = {r.get("品名", "未定義"): {"大カテゴリ": r.get("大カテゴリ", "その他"), "中カテゴリ": r.get("中カテゴリ", "その他"), "成分": safe_parse_recipe(r.get("配合JSON"))} for r in recipes_raw if r.get("大カテゴリ") != "調味料"}
    seasoning_recipes_all = [r for r in recipes_raw if r.get("大カテゴリ") == "調味料"]

    sec_title("🏭 ライン・製品選択")
    st.markdown('<div style="font-weight:800; color:#64748b; margin-bottom:8px;">① ラインを選択</div>', unsafe_allow_html=True)
    BASE_BIG_CAT_ORDER = ["プラント", "OKM", "手詰め"]
    dynamic_cats = {v["大カテゴリ"] for v in p_recipes.values() if v.get("大カテゴリ")}
    big_cats = list(BASE_BIG_CAT_ORDER) + sorted(dynamic_cats - set(BASE_BIG_CAT_ORDER))
    big_cat_labels = [f"{big_cat_icon(c)} {c}" for c in big_cats]
    sel_big_label = st.radio("ライン", big_cat_labels, horizontal=True, label_visibility="collapsed") if big_cats else None
    big_cat = big_cats[big_cat_labels.index(sel_big_label)] if sel_big_label else None

    SUB_CAT_ORDER = ["黒", "白", "耐冷", "ショクカイ", "めん", "その他"]
    sub_cats_set = {v["中カテゴリ"] for v in p_recipes.values() if v.get("大カテゴリ") == big_cat and v.get("中カテゴリ")} if big_cat else set()
    sub_cats = sorted(sub_cats_set, key=lambda c: (SUB_CAT_ORDER.index(c) if c in SUB_CAT_ORDER else len(SUB_CAT_ORDER), c))
    sub_str = None
    if big_cat and len(sub_cats) > 1:
        st.markdown('<div style="font-weight:800; color:#64748b; margin:24px 0 8px 0;">② 種別を選択</div>', unsafe_allow_html=True)
        sub_cat_labels = [f"{sub_cat_icon(c)} {c}" for c in sub_cats]
        sel_sub_label = st.radio("種別", sub_cat_labels, horizontal=True, label_visibility="collapsed")
        sub_str = sub_cats[sub_cat_labels.index(sel_sub_label)]
    elif sub_cats: sub_str = sub_cats[0]

    st.markdown('<div style="font-weight:800; color:#64748b; margin:24px 0 8px 0;">③ 製品品番を選択</div>', unsafe_allow_html=True)
    filtered_opts = [k for k, v in p_recipes.items() if v.get("大カテゴリ") == big_cat and v.get("中カテゴリ") == sub_str] if big_cat and sub_str else []
    selected_p = None
    active_recipe = []
    if filtered_opts:
        opt_labels = [f"{product_icon(k)} {k}" for k in filtered_opts]
        sel_label = st.radio("製品", opt_labels, horizontal=True, label_visibility="collapsed")
        selected_p = filtered_opts[opt_labels.index(sel_label)]
        active_recipe = p_recipes.get(selected_p, {}).get("成分", [])
    card_end()

    if not active_recipe: st.info("👆 製品を選択してください。")
    else:
        card_start()
        sec_title("⚖️ 希望仕込量と石灰水量")

        if "t_size" not in st.session_state: st.session_state["t_size"] = 100.0
        if "l_size" not in st.session_state: st.session_state["l_size"] = 0.0
        def add_t_size(v): st.session_state["t_size"] = max(0.0, st.session_state["t_size"] + v)
        def add_l_size(v): st.session_state["l_size"] = max(0.0, st.session_state["l_size"] + v)

        col_in1, col_in2 = st.columns(2)
        with col_in1:
            st.markdown("<div style='font-weight:800; color:#475569; margin-bottom:6px;'>🏭 希望仕込製品量 (kg)</div>", unsafe_allow_html=True)
            c1, c2, c3, c4 = st.columns(4)
            c1.button("+1000", key="btn_t_1000", on_click=add_t_size, args=(1000,), use_container_width=True)
            c2.button("+100",  key="btn_t_100",  on_click=add_t_size, args=(100,),  use_container_width=True)
            c3.button("+10",   key="btn_t_10",   on_click=add_t_size, args=(10,),   use_container_width=True)
            c4.button("✖0",    key="btn_t_0",    on_click=lambda: st.session_state.update({"t_size": 0.0}), use_container_width=True)
            target_size = st.number_input("仕込量", min_value=0.0, step=10.0, key="t_size", label_visibility="collapsed", format="%.0f")

        with col_in2:
            st.markdown("<div style='font-weight:800; color:#475569; margin-bottom:6px;'>💧 石灰水作成量 (kg)</div>", unsafe_allow_html=True)
            c1, c2, c3, c4 = st.columns(4)
            c1.button("+100", key="btn_l_100", on_click=add_l_size, args=(100,), use_container_width=True)
            c2.button("+10",  key="btn_l_10",  on_click=add_l_size, args=(10,),  use_container_width=True)
            c3.button("+1",   key="btn_l_1",   on_click=add_l_size, args=(1,),   use_container_width=True)
            c4.button("✖0",   key="btn_l_0",   on_click=lambda: st.session_state.update({"l_size": 0.0}), use_container_width=True)
            lime_water_size = st.number_input("石灰水量", min_value=0.0, step=1.0, key="l_size", label_visibility="collapsed")
        
        st.markdown("<br>", unsafe_allow_html=True)
        c_op1, c_op2 = st.columns(2)
        with c_op1: operator = render_operator_selector("op_key")
        with c_op2: brew_remarks = st.text_input("📝 備考（任意）", placeholder="特記事項があれば入力")
        card_end()

        if target_size > 0:
            st.markdown('<div class="section-title" style="margin-top:32px;">📦 準備する原料・ロット</div>', unsafe_allow_html=True)
            submitted_ingredients, water_items = [], []
            lime_cfg = parse_lime_config(order_points, product_name=selected_p)
            lime_boost_active = is_lime_boost_active(lime_cfg, brew_date)

            for i, item in enumerate(active_recipe[:15]):
                r_name = str(item.get("原料名", "")).strip()
                base_ratio = float(item.get("比率", 0.0))
                is_water, is_lime, is_konjac = ("水" in r_name or "お湯" in r_name), ("石灰" in r_name or "カルシウム" in r_name), ("こんにゃく" in r_name)
                icon = "🧂" if is_lime else ("📦" if is_konjac else "🔹")

                lime_msg = ""
                if is_water: 
                    calc_kg = max(0.0, target_size * (base_ratio / 100.0) - lime_water_size)
                elif is_lime: 
                    eff_ratio = base_ratio
                    if lime_boost_active:
                        add_r = float(lime_cfg.get("add_ratio", 0.01))
                        eff_ratio += add_r
                        lime_msg = f"🌡️ 期間増量中 ({lime_cfg.get('start_month', 6)}月〜{lime_cfg.get('end_month', 9)}月: +{add_r}% / 理由: {lime_cfg.get('reason', '')})"
                    calc_kg = lime_water_size * (eff_ratio / 10.0)
                else: calc_kg = target_size * (base_ratio / 100.0)

                if is_water:
                    water_items.append({"原料名": r_name, "kg": round(calc_kg, 2), "配合比": base_ratio})
                    submitted_ingredients.append({"原料名": r_name, "kg": round(calc_kg, 2), "lot": "─"})
                    continue

                with st.container(border=True):
                    st.markdown(f"<div style='font-size:1.3rem; font-weight:900;'>{icon} {r_name}</div>", unsafe_allow_html=True)
                    if is_lime and lime_msg: st.markdown(f"<div style='font-size:0.9rem; color:#b45309; font-weight:800; margin-top:4px;'>{lime_msg}</div>", unsafe_allow_html=True)

                    if is_konjac:
                        st.markdown("<div style='background:#f8fafc; padding:12px; border-radius:8px; border:1px solid #e2e8f0; margin-top:8px;'>", unsafe_allow_html=True)
                        blend_mode = st.radio("ブレンドモード", ["単一 (1種)", "2種ブレンド", "3種ブレンド"], key=f"kb_{selected_p}_{i}", horizontal=True, label_visibility="collapsed")
                        konjac_mats = [m for m in materials if "こんにゃく" in m] or [r_name]
                        
                        if blend_mode in ["2種ブレンド", "3種ブレンド"]:
                            is_3 = (blend_mode == "3種ブレンド")
                            k_a, k_b, k_c = f"kr_a_{selected_p}_{i}", f"kr_b_{selected_p}_{i}", f"kr_c_{selected_p}_{i}"
                            if k_a not in st.session_state: st.session_state[k_a] = 50.0 if not is_3 else 34.0
                            if k_b not in st.session_state: st.session_state[k_b] = 50.0 if not is_3 else 33.0
                            if k_c not in st.session_state: st.session_state[k_c] = 33.0

                            st.markdown("<div style='background:#ffffff; padding:12px 16px; border-radius:8px; border:1px solid #cbd5e1; margin-bottom:16px;'>", unsafe_allow_html=True)
                            st.markdown("<div style='font-size:0.9rem; font-weight:800; color:#475569; margin-bottom:8px;'>🎯 ワンタッチ比率入力パネル</div>", unsafe_allow_html=True)
                            
                            target_key = f"blend_tgt_{selected_p}_{i}"
                            if target_key not in st.session_state: st.session_state[target_key] = "🅰️"
                            st.radio("入力対象を選択", ["🅰️", "🅱️", "🅲"] if is_3 else ["🅰️", "🅱️"], horizontal=True, key=target_key, label_visibility="collapsed")
                            
                            st.markdown("<div class='ratio-btn-container' style='margin-top:8px;'>", unsafe_allow_html=True)
                            btn_cols = st.columns(9)
                            def update_ratio(v, tgt, ka, kb, kc, is_three):
                                if tgt == "🅰️": st.session_state[ka] = float(v); st.session_state[kb] = 100.0 - float(v) if not is_three else st.session_state[kb]
                                elif tgt == "🅱️": st.session_state[kb] = float(v); st.session_state[ka] = 100.0 - float(v) if not is_three else st.session_state[ka]
                                elif tgt == "🅲" and is_three: st.session_state[kc] = float(v)
                            
                            for pidx, pv in enumerate(range(10, 100, 10)):
                                curr_tgt = st.session_state[target_key]
                                curr_val = st.session_state.get(k_a) if curr_tgt == "🅰️" else (st.session_state.get(k_b) if curr_tgt == "🅱️" else st.session_state.get(k_c))
                                btn_cols[pidx].button(f"{pv}%", key=f"rbtn_{selected_p}_{i}_{pv}", on_click=update_ratio, args=(pv, curr_tgt, k_a, k_b, k_c, is_3), type="primary" if (curr_val == float(pv)) else "secondary", use_container_width=True)
                            st.markdown("</div>", unsafe_allow_html=True)

                            cols_ratio = st.columns(3 if is_3 else 2)
                            ratio_a = cols_ratio[0].number_input("🅰️ 比率(%)", min_value=0.0, max_value=100.0, step=1.0, key=k_a)
                            ratio_b = cols_ratio[1].number_input("🅱️ 比率(%)", min_value=0.0, max_value=100.0, step=1.0, key=k_b)
                            ratio_c = cols_ratio[2].number_input("🅲 比率(%)", min_value=0.0, max_value=100.0, step=1.0, key=k_c) if is_3 else 0.0
                            st.markdown("</div>", unsafe_allow_html=True)

                            mat_a = st.radio("🅰️ 原料種別", konjac_mats, key=f"kma_{selected_p}_{i}", horizontal=True, label_visibility="collapsed")
                            ca_amt, ca_lot = st.columns([1, 1])
                            with ca_amt: act_a = render_amount_adjuster(f"🅰️ 投入量", calc_kg * ratio_a / 100.0, f"adj_a_{selected_p}_{i}")
                            with ca_lot: lot_a = render_lot_selector(mat_a, f"lot_a_{selected_p}_{i}_{mat_a}")
                            st.markdown("<hr style='margin:16px 0; border-top:1px dashed #cbd5e1;'>", unsafe_allow_html=True)
                            
                            mat_b = st.radio("🅱️ 原料種別", konjac_mats, index=1 if len(konjac_mats)>1 else 0, key=f"kmb_{selected_p}_{i}", horizontal=True, label_visibility="collapsed")
                            cb_amt, cb_lot = st.columns([1, 1])
                            with cb_amt: act_b = render_amount_adjuster(f"🅱️ 投入量", calc_kg * ratio_b / 100.0, f"adj_b_{selected_p}_{i}")
                            with cb_lot: lot_b = render_lot_selector(mat_b, f"lot_b_{selected_p}_{i}_{mat_b}")

                            submitted_ingredients.extend([{"原料名": mat_a, "kg": act_a, "lot": f"{lot_a}({ratio_a}%)"}, {"原料名": mat_b, "kg": act_b, "lot": f"{lot_b}({ratio_b}%)"}])
                            
                            if is_3:
                                st.markdown("<hr style='margin:16px 0; border-top:1px dashed #cbd5e1;'>", unsafe_allow_html=True)
                                mat_c = st.radio("🅲 原料種別", konjac_mats, index=2 if len(konjac_mats)>2 else 0, key=f"kmc_{selected_p}_{i}", horizontal=True, label_visibility="collapsed")
                                cc_amt, cc_lot = st.columns([1, 1])
                                with cc_amt: act_c = render_amount_adjuster(f"🅲 投入量", calc_kg * ratio_c / 100.0, f"adj_c_{selected_p}_{i}")
                                with cc_lot: lot_c = render_lot_selector(mat_c, f"lot_c_{selected_p}_{i}_{mat_c}")
                                submitted_ingredients.append({"原料名": mat_c, "kg": act_c, "lot": f"{lot_c}({ratio_c}%)"})
                        else:
                            c_amt, c_lot = st.columns([1, 1])
                            with c_amt: act_kg = render_amount_adjuster(f"投入量（配合比 {fmt_kg(base_ratio)}%）", calc_kg, f"adj_{selected_p}_{i}")
                            with c_lot: final_lot = render_lot_selector(r_name, f"lot_{selected_p}_{i}")
                            submitted_ingredients.append({"原料名": r_name, "kg": act_kg, "lot": final_lot})
                        st.markdown("</div>", unsafe_allow_html=True)
                    else:
                        c_amt, c_lot = st.columns([1, 1])
                        with c_amt: act_kg = render_amount_adjuster(f"投入量（配合比 {fmt_kg(base_ratio)}%）", calc_kg, f"adj_{selected_p}_{i}")
                        with c_lot: final_lot = render_lot_selector(r_name, f"lot_{selected_p}_{i}")
                        submitted_ingredients.append({"原料名": r_name, "kg": act_kg, "lot": final_lot})

            if seasoning_recipes_all:
                st.markdown('<div class="section-title" style="margin-top:32px;">🌶️ 調味料の希釈計算・投入記録</div>', unsafe_allow_html=True)
                for sr_idx, sr in enumerate(seasoning_recipes_all):
                    sr_items = safe_parse_seasoning_recipe(sr.get("配合JSON"))
                    if not sr_items: continue
                    with st.container(border=True):
                        if st.checkbox(f"🌶️ {sr.get('品名')} を使用する", key=f"use_season_{selected_p}_{sr_idx}"):
                            target_vol_key = f"season_vol_{selected_p}_{sr_idx}"
                            target_vol = st.number_input("希釈対象量（できあがり量・kg）", min_value=0.0, value=st.session_state.get(target_vol_key, 0.0), step=0.1, key=target_vol_key)
                            for si, sitem in enumerate(sr_items):
                                s_mat, s_dil = sitem["原料名"], sitem["希釈倍率"]
                                need_kg = target_vol / s_dil if s_dil > 0 else 0.0
                                sc_amt, sc_lot = st.columns([1, 1])
                                with sc_amt: s_act_kg = render_amount_adjuster(f"投入量（{s_mat}）", need_kg, f"adj_season_{selected_p}_{sr_idx}_{si}")
                                with sc_lot: s_lot = render_lot_selector(s_mat, f"lot_season_{selected_p}_{sr_idx}_{si}")
                                submitted_ingredients.append({"原料名": s_mat, "kg": s_act_kg, "lot": s_lot})

            st.markdown("<br>", unsafe_allow_html=True)
            total_in = sum(ing["kg"] for ing in submitted_ingredients)
            st.markdown(f"""
            <div style="background-color: var(--c-primary-soft); border: 2px solid var(--c-primary); border-radius: 12px; padding: 16px; margin-bottom: 24px; text-align: center;">
                <div style="font-weight: 800; color: var(--c-muted); font-size: 1.1rem;">💡 合計投入予定量（全原料・水を含む）</div>
                <div style="font-size: 2.2rem; font-weight: 900; color: var(--c-secondary);">{fmt_kg(total_in)} <span style="font-size:1.2rem; color:var(--c-muted);">kg</span></div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("💾 この内容で製造記録を保存する", type="primary", use_container_width=True):
                k_kg = s_kg = st_kg = lime_kg = 0.0
                k_lot = s_lot = st_lot = "─"
                for ing in submitted_ingredients:
                    n, amt, lot = ing["原料名"], ing["kg"], ing["lot"]
                    if "こんにゃく" in n: k_kg += amt; k_lot = lot if k_lot == "─" else (k_lot if lot in k_lot else f"{k_lot} / {lot}")
                    elif "海藻" in n: s_kg += amt; s_lot = lot if s_lot == "─" else (s_lot if lot in s_lot else f"{s_lot} / {lot}")
                    elif "デンプン" in n or "でんぷん" in n: st_kg += amt; st_lot = lot if st_lot == "─" else (st_lot if lot in st_lot else f"{st_lot} / {lot}")
                    elif "石灰" in n or "カルシウム" in n: lime_kg += amt

                next_no = sheets.next_brewing_no(brewing) if hasattr(sheets, "next_brewing_no") else f"BRW-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                text_ing = ", ".join([f"{ing['原料名']}:{ing['kg']}kg({ing['lot']})" for ing in submitted_ingredients])

                if hasattr(sheets, "append_brewing"):
                    sheets.append_brewing({
                        "仕込No": next_no, "仕込日": str(brew_date), "品名": selected_p,
                        "メーカー": operator, "主原料ロット": k_lot, "仕込量(kg)": round(target_size, 2),
                        "こんにゃく精粉(kg)": round(k_kg, 2), "海藻粉(kg)": round(s_kg, 2), "海藻粉ロット": s_lot,
                        "デンプン(kg)": round(st_kg, 2), "デンプンロット": st_lot, "デンプン種別": "-",
                        "石灰(kg)": round(lime_kg, 2), "石灰水(L)": round(lime_water_size, 2),
                        "その他添加物": text_ing, "備考": f"{brew_remarks}", "登録日時": datetime.now().isoformat()
                    })
                
                for key in list(st.session_state.keys()):
                    if any(key.startswith(p) for p in ["adj_", "last_calc_", "lot_", "btn_", "kb_", "kr_", "km", "blend_tgt_", "use_season_", "season_vol_"]):
                        del st.session_state[key]
                
                st.toast("✅ 製造記録を保存しました", icon="💾")
                st.markdown(f'<div style="background-color: #dcfce7; border: 2px solid #22c55e; border-radius: 12px; padding: 18px; margin-top: 16px; text-align: center;"><div style="font-size: 1.4rem; font-weight: 900; color: #15803d;">✅ 製造記録を正しく登録しました (仕込No. {next_no})</div></div>', unsafe_allow_html=True)
                time.sleep(1.8); refresh()


# ═══════════════════════════════════════════════════════════════
#  📊 ダッシュボード
# ═══════════════════════════════════════════════════════════════
elif page == "📊 ダッシュボード":
    st.markdown('<div class="main-header"><h1>📊 サマリーと在庫モニター</h1></div>', unsafe_allow_html=True)
    df_brw_global = pd.DataFrame(brewing)
    if not df_brw_global.empty:
        df_brw_global["仕込日_dt"] = pd.to_datetime(df_brw_global["仕込日"], errors="coerce")
        df_brw_today = df_brw_global[df_brw_global["仕込日_dt"].dt.strftime("%Y-%m-%d") == date.today().strftime("%Y-%m-%d")]
        today_total_kg = pd.to_numeric(df_brw_today["仕込量(kg)"], errors="coerce").fillna(0).sum()
        today_count = len(df_brw_today)
    else: today_total_kg = today_count = 0

    c1, c2, c3 = st.columns(3)
    c1.metric("📦 本日の総製造量", f"{fmt_kg(today_total_kg)} kg", f"{today_count} 件製造")
    c2.metric("⚠️ 在庫不足原料", f"{sum(1 for m in materials if parse_op_data(order_points.get(m, 0.0))[0] > 0 and type_totals_bag.get(m, 0.0) < parse_op_data(order_points.get(m, 0.0))[0])} 品目")
    po_all = parse_purchase_orders(order_points)
    po_pending = [o for o in po_all if o.get("ステータス") != "入荷済み"]
    po_overdue = sum(1 for o in po_pending if o.get("納品予定日") and o["納品予定日"] < str(date.today()))
    c3.metric("📝 未入荷の発注", f"{len(po_pending)} 件", f"うち超過 {po_overdue} 件" if po_overdue else None, delta_color="inverse")

    st.markdown("---")
    sec_title("📦 主要原料 現在庫とアラート")
    st.caption("※ 棚卸し等の在庫調整は「📦 在庫・棚卸」ページから行ってください。")
    cols = st.columns(min(3, len(materials) if materials else 1))
    
    for idx, m in enumerate(materials):
        pt, wt = parse_op_data(order_points.get(m, 0.0))
        curr_kg = type_totals_kg.get(m, 0.0)
        curr_bag = curr_kg / wt if wt > 0 else 0
        is_alert = (pt > 0 and curr_bag < pt)
        bg_col, b_col = ("#fef2f2", "#ef4444") if is_alert else ("#ffffff", "#cbd5e1")
        alert_msg = f"<div style='font-size:0.9rem; color:#ef4444; font-weight:bold; margin-top:8px;'>⚠️ 発注点({fmt_kg(pt)}袋) 以下</div>" if is_alert else f"<div style='font-size:0.9rem; color:#64748b; font-weight:bold; margin-top:8px;'>✅ 発注点: {fmt_kg(pt)}袋</div>"

        img_b64 = get_material_image(order_points, m)
        img_html = f'<img src="{img_b64}" style="width:50px; height:50px; object-fit:cover; border-radius:6px; margin-right:12px; border:1px solid #e2e8f0;">' if img_b64 else f'<div style="width:50px; height:50px; background:#e2e8f0; border-radius:6px; margin-right:12px; display:flex; align-items:center; justify-content:center; font-size:20px;">🥦</div>'

        with cols[idx % 3]:
            st.markdown(f"""
            <div style="background:{bg_col}; border:2px solid {b_col}; border-radius:12px; padding:18px; margin-bottom:8px;">
                <div style="display:flex; align-items:center; margin-bottom:8px;">
                    {img_html}
                    <div style="font-weight:900; color:#0f172a; font-size:1.15rem;">{m}</div>
                </div>
                <div class="mat-card-value" style="font-size:2.2rem; font-weight:900; color:#0f766e; margin:6px 0 2px 0;">
                    {fmt_kg(curr_kg)}<span style="font-size:1.1rem; color:#64748b; margin-right:8px;">kg</span> 
                    <span style="font-size:1.6rem; color:#0f172a;">({fmt_kg(curr_bag)}袋)</span>
                </div>
                <div style="font-size:0.85rem; color:#64748b; margin-bottom:4px;">1袋 = {fmt_kg(wt)} kg 換算</div>
                {alert_msg}
            </div>
            """, unsafe_allow_html=True)

            mat_lots = get_lots_for_material(m)
            with st.popover(f"🔧 {m} を増減・棚卸調整", use_container_width=True):
                if not mat_lots: st.info("この原料の入荷記録がありません。")
                else:
                    st.markdown(f"**🔧 {m} の在庫調整**")
                    lot_opts = {f"{v['ロットNo']} (現在庫:{fmt_kg(v['現在庫(袋)'])}袋)": v["入荷No"] for v in mat_lots}
                    sel_lot = st.selectbox("対象ロット", list(lot_opts.keys()), key=f"dash_lot_{m}")
                    target_ano = lot_opts[sel_lot]
                    target_lot_data = next(v for v in mat_lots if v["入荷No"] == target_ano)
                    
                    adj_mode = st.radio("調整方法", ["➕➖ クイック増減", "📋 実地数量で確定"], horizontal=True, key=f"dash_mode_{m}")
                    op_q = render_operator_selector(f"dash_qop_{m}")
                    
                    def _dash_adj(ano, delta_bags, reason_text):
                        if hasattr(sheets, "append_adjustment"):
                            sheets.append_adjustment({
                                "調整ID": f"ADJ-{datetime.now().strftime('%Y%m%d%H%M%S%f')}", "入荷No": ano,
                                "調整日": str(date.today()), "調整袋数": delta_bags, "理由": reason_text,
                                "担当者": op_q, "登録日時": datetime.now().isoformat()
                            })
                            
                    if "クイック" in adj_mode:
                        st.caption("ボタンをタップした瞬間に在庫が即時増減します")
                        qc1, qc2, qc3, qc4 = st.columns(4)
                        if qc1.button("➖10", key=f"dq_m10_{m}", use_container_width=True): _dash_adj(target_ano, -10, "【クイック増減:-10】"); st.toast("-10袋"); time.sleep(1); refresh()
                        if qc2.button("➖1", key=f"dq_m1_{m}", use_container_width=True): _dash_adj(target_ano, -1, "【クイック増減:-1】"); st.toast("-1袋"); time.sleep(1); refresh()
                        if qc3.button("➕1", key=f"dq_p1_{m}", use_container_width=True): _dash_adj(target_ano, 1, "【クイック増減:+1】"); st.toast("+1袋"); time.sleep(1); refresh()
                        if qc4.button("➕10", key=f"dq_p10_{m}", use_container_width=True): _dash_adj(target_ano, 10, "【クイック増減:+10】"); st.toast("+10袋"); time.sleep(1); refresh()
                    else:
                        st.caption("実際に数えた袋数を入力してください。差分が自動記録されます（0にすれば在庫ゼロになります）。")
                        actual = st.number_input("実在庫数量(袋)", min_value=0.0, value=float(round(target_lot_data["現在庫(袋)"], 2)), step=1.0, key=f"dash_act_{m}")
                        diff = round(actual - target_lot_data["現在庫(袋)"], 4)
                        if st.button("💾 この実地数量で確定", type="primary", key=f"dash_save_{m}", use_container_width=True):
                            _dash_adj(target_ano, diff, f"【実地棚卸で {fmt_kg(actual)}袋 に更新】")
                            st.success(f"{fmt_kg(actual)}袋で確定しました"); time.sleep(1.5); refresh()

            if is_konjac_material(m) and mat_lots:
                breakdown = {}
                for v in mat_lots:
                    breakdown.setdefault((v["メーカー"], v["グレード"]), {"kg": 0.0, "bag": 0.0})
                    breakdown[(v["メーカー"], v["グレード"])]["kg"] += v["現在庫(kg)"]
                    breakdown[(v["メーカー"], v["グレード"])]["bag"] += v["現在庫(袋)"]
                with st.expander(f"🏷️ {m} のメーカー・グレード別内訳"):
                    for (mk, gr), vals in sorted(breakdown.items()):
                        if vals["kg"] > 0.01:
                            st.markdown(f'<div style="display:flex; justify-content:space-between; align-items:center; padding:8px 10px; border-bottom:1px solid #e2e8f0;"><div style="font-weight:800; font-size:0.85rem;">🏢 {mk} / 🏷️ {gr}</div><div style="font-weight:900; color:#0f766e;">{fmt_kg(vals["kg"])} kg（{fmt_kg(vals["bag"])}袋）</div></div>', unsafe_allow_html=True)
            st.markdown("<div style='margin-bottom:16px;'></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  📝 発注管理
# ═══════════════════════════════════════════════════════════════
elif page == "📝 発注管理":
    st.markdown('<div class="main-header"><h1>📝 原料・資材 発注管理</h1></div>', unsafe_allow_html=True)
    all_orders = parse_purchase_orders(order_points)
    pending_orders = [o for o in all_orders if o.get("ステータス") != "入荷済み"]
    done_orders = [o for o in all_orders if o.get("ステータス") == "入荷済み"]

    t_new, t_list = st.tabs(["➕ 新規発注登録", "📋 発注一覧・入荷処理"])
    with t_new:
        card_start()
        order_category = st.radio("🏷️ 発注区分", ["🥦 原材料", "🧻 衛生資材・消耗品"], horizontal=True)
        is_raw_mat = "原材料" in order_category
        item_opts = materials if is_raw_mat else [s.get("資材名") for s in supplies]
        with st.form("new_order_form"):
            o_mat = st.selectbox("発注品目", item_opts if item_opts else ["未登録"])
            o_maker = st.selectbox("メーカー/発注先", makers if makers else ["未登録"])
            c_a, c_b = st.columns(2)
            o_date = c_a.date_input("発注日", value=date.today())
            o_due = c_b.date_input("納品予定日", value=date.today() + timedelta(days=7))
            o_qty = st.number_input("発注個数 (袋/箱など)", min_value=1, value=10, step=1)
            o_note = st.text_input("備考（任意）")
            if st.form_submit_button("💾 発注を登録する", type="primary", use_container_width=True):
                all_orders.append({
                    "発注ID": f"PO-{datetime.now().strftime('%Y%m%d%H%M%S%f')}", "発注区分": "原材料" if is_raw_mat else "衛生資材",
                    "発注日": str(o_date), "原料名": o_mat, "メーカー": o_maker, "個数": o_qty, "納品予定日": str(o_due), "ステータス": "未入荷",
                    "紐づく入荷No": "", "備考": o_note, "登録日時": datetime.now().isoformat()
                })
                save_purchase_orders(order_points, all_orders)
                st.success("発注を登録しました。"); time.sleep(1); refresh()
        card_end()

    with t_list:
        if pending_orders:
            for o in sorted(pending_orders, key=lambda x: x.get("納品予定日", "")):
                is_raw_order = (o.get("発注区分", "原材料") == "原材料")
                with st.container(border=True):
                    c_i, c_b = st.columns([3, 1])
                    with c_i:
                        st.markdown(f"**{'🥦' if is_raw_order else '🧻'} {o.get('原料名')}**　🏢 {o.get('メーカー')}　📦 {fmt_kg(o.get('個数'))}")
                        st.caption(f"発注日: {o.get('発注日')} ／ 納品予定日: {o.get('納品予定日')}")
                    with c_b:
                        with st.popover("✅ 入荷処理"):
                            oid = o.get("発注ID")
                            st.markdown(f"#### ✅ {o.get('原料名')} の入荷")
                            if is_raw_order:
                                arr_lot = st.text_input("ロットNo ＊必須", key=f"po_lot_{oid}")
                                grade_list = parse_grade_list(order_points)
                                po_grade = st.selectbox("🏷️ グレード", grade_list, key=f"po_grade_{oid}") if is_konjac_material(o.get("原料名")) and grade_list else "-"
                                _, po_default_wt = parse_op_data(order_points.get(o.get("原料名"), 0))
                                pc1, pc2 = st.columns(2)
                                po_bags = pc1.number_input("入荷袋数", min_value=1, value=int(float(o.get("個数", 1))), step=1, key=f"po_bags_{oid}")
                                po_wpb = pc2.number_input("1袋重量(kg)", min_value=1, value=int(float(po_default_wt)), step=1, key=f"po_wpb_{oid}")
                                
                                st.markdown("**🔍 受入品質検査**")
                                po_chk = {k: st.radio(v, ["未確認", "✅ 正常", "❌ 異常あり"], horizontal=True, key=f"po_chk_{oid}_{k}") for k, v in [("外観", "📦 外観"), ("品名", "🏷️ 品名"), ("賞味期限", "📅 賞味期限"), ("異物", "🔍 異物")]}
                                po_op = render_operator_selector(f"po_op_{oid}")
                                if st.button("💾 入荷を登録", type="primary", use_container_width=True, key=f"po_save_{oid}"):
                                    if not arr_lot: st.error("ロットNo必須です")
                                    elif any(v=="未確認" for v in po_chk.values()): st.error("検査未完了です")
                                    elif hasattr(sheets, "append_arrival"):
                                        new_ano = sheets.next_arrival_no(arrivals)
                                        sheets.append_arrival({
                                            "入荷No": new_ano, "入荷日": str(date.today()), "メーカー": o.get("メーカー"), "ロットNo": arr_lot,
                                            "原料種別": o.get("原料名"), "グレード": po_grade, "袋数": po_bags, "1袋重量(kg)": po_wpb, "総量(kg)": po_bags * po_wpb,
                                            "外観": po_chk["外観"], "品名・規格確認": po_chk["品名"], "賞味期限": po_chk["賞味期限"], "異物": po_chk["異物"],
                                            "担当者": po_op, "備考": f"発注ID:{oid}", "登録日時": datetime.now().isoformat()
                                        })
                                        for oo in all_orders:
                                            if oo.get("発注ID") == oid: oo.update({"ステータス": "入荷済み", "紐づく入荷No": new_ano, "入荷処理日": str(date.today())})
                                        save_purchase_orders(order_points, all_orders)
                                        st.success("入荷しました。"); time.sleep(1.5); refresh()
                            else:
                                po_bags = st.number_input("入荷個数", min_value=1, value=int(float(o.get("個数", 1))), step=1, key=f"po_bags_sup_{oid}")
                                po_op = render_operator_selector(f"po_op_sup_{oid}")
                                if st.button("💾 資材在庫に追加", type="primary", use_container_width=True, key=f"po_save_sup_{oid}"):
                                    sid = next((s.get("資材ID") for s in supplies if s.get("資材名") == o.get("原料名")), None)
                                    if sid and hasattr(sheets, "append_supply_log"):
                                        sheets.append_supply_log({
                                            "ログID": f"LOG-{datetime.now().strftime('%Y%m%d%H%M%S%f')}", "登録日": str(date.today()), "資材ID": sid, 
                                            "処理": "入荷", "数量": po_bags, "作業者": po_op, "備考": f"発注:{oid}", "登録日時": datetime.now().isoformat()
                                        })
                                        for oo in all_orders:
                                            if oo.get("発注ID") == oid: oo.update({"ステータス": "入荷済み", "紐づく入荷No": "資材入荷", "入荷処理日": str(date.today())})
                                        save_purchase_orders(order_points, all_orders)
                                        st.success("資材を入荷しました。"); time.sleep(1.5); refresh()

        if done_orders:
            sec_title("✅ 入荷済みの発注（履歴）")
            df_done = pd.DataFrame(sorted(done_orders, key=lambda x: x.get("入荷処理日", ""), reverse=True))
            st.dataframe(fmt_df_numeric(df_done[[c for c in ["発注区分", "発注日", "原料名", "メーカー", "個数", "入荷処理日", "紐づく入荷No"] if c in df_done.columns]].head(50), ["個数"]), use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════
#  📥 入荷登録
# ═══════════════════════════════════════════════════════════════
elif page == "📥 入荷登録":
    st.markdown('<div class="main-header"><h1>📥 原料入荷品質記録</h1></div>', unsafe_allow_html=True)
    def _save_arrivals_recalc(records):
        for r in records:
            try: r["総量(kg)"] = float(r.get("袋数", 0) or 0) * float(r.get("1袋重量(kg)", 0) or 0)
            except: pass
        if hasattr(sheets, "save_arrivals"): sheets.save_arrivals(records)
            
    t_in, t_hist = st.tabs(["➕ 新規入荷登録", "📋 入荷履歴・編集"])
    with t_in:
        card_start()
        new_no = sheets.next_arrival_no(arrivals) if hasattr(sheets, "next_arrival_no") else ""
        arr_date = st.date_input("入荷日", value=date.today())
        maker_sel = st.selectbox("メーカー", makers if makers else ["未登録"])
        lot_val = st.text_input("ロットNo ＊必須", placeholder="例: L12345")
        
        m_type = st.selectbox("原料種別", materials if materials else ["未登録"])
        _, default_wt = parse_op_data(order_points.get(m_type, 0))
        grade_val = st.selectbox("🏷️ グレード", parse_grade_list(order_points)) if is_konjac_material(m_type) and parse_grade_list(order_points) else "-"
        
        c1, c2 = st.columns(2)
        bags_qty = c1.number_input("入荷袋数", min_value=1, value=10, step=1)
        weight_per_bag = c2.number_input("1袋重量 (kg) ※自動セット済", min_value=1, value=int(float(default_wt)), step=1)
        
        sec_title("🔍 受入品質検査")
        chk_results = {}
        cols_chk = st.columns(2)
        for idx, (k_n, lbl) in enumerate([("外観", "📦 外観"), ("品名・規格確認", "🏷️ 品名・規格"), ("賞味期限", "📅 賞味期限"), ("異物", "🔍 異物混入")]):
            with cols_chk[idx % 2]:
                with st.container(border=True):
                    st.markdown(f"**{lbl}**")
                    chk_results[k_n] = st.radio(lbl, ["未確認", "✅ 正常", "❌ 異常あり"], index=0, key=f"chk_{k_n}", horizontal=True, label_visibility="collapsed")
        
        chk_note = st.text_input("備考")
        operator = render_operator_selector("arr_op")
        
        if st.button("💾 入荷記録を登録する", type="primary", use_container_width=True):
            if not lot_val: st.error("ロットNoは必須項目です。")
            elif any(v=="未確認" for v in chk_results.values()): st.error("受入品質検査が未完了です。")
            elif hasattr(sheets, "append_arrival"):
                sheets.append_arrival({
                    "入荷No": new_no, "入荷日": str(arr_date), "メーカー": maker_sel, "ロットNo": lot_val,
                    "原料種別": m_type, "グレード": grade_val, "袋数": bags_qty, "1袋重量(kg)": weight_per_bag, "総量(kg)": bags_qty * weight_per_bag,
                    "外観": chk_results["外観"], "品名・規格確認": chk_results["品名・規格確認"], "賞味期限": chk_results["賞味期限"], "異物": chk_results["異物"],
                    "担当者": operator, "備考": chk_note, "登録日時": datetime.now().isoformat()
                })
                st.success("入荷記録を保存しました。"); time.sleep(1.5); refresh()
        card_end()
        
    with t_hist:
        if arrivals:
            df_arr = pd.DataFrame(arrivals).sort_values("入荷日", ascending=False).reset_index(drop=True)
            card_start()
            sec_title("✏️ 入荷履歴の一括編集（Excel風）")
            arr_editable_cols = ["入荷日", "原料種別", "メーカー", "ロットNo", "袋数", "1袋重量(kg)", "備考"]
            if "グレード" in df_arr.columns: arr_editable_cols.insert(2, "グレード")
            render_excel_history_editor(full_records=arrivals, filtered_df=df_arr.head(200), id_col="入荷No", editable_cols=arr_editable_cols, numeric_cols=["袋数", "1袋重量(kg)"], save_func=_save_arrivals_recalc, key_prefix="arr_hist")
            card_end()


# ═══════════════════════════════════════════════════════════════
#  📦 在庫・棚卸
# ═══════════════════════════════════════════════════════════════
elif page == "📦 在庫・棚卸":
    st.markdown('<div class="main-header"><h1>📦 在庫・棚卸管理</h1></div>', unsafe_allow_html=True)
    t_inv, t_hist = st.tabs(["📋 在庫一覧・棚卸し", "🕒 変更履歴"])
    
    with t_inv:
        card_start()
        active_inv = [v for v in inventory_data.values() if v["現在庫(袋)"] > 0.001]
        if active_inv: render_lot_inventory_manager(active_inv)
        else: st.info("在庫データがありません。")
        card_end()

    with t_hist:
        if adjustments:
            df_adj = pd.DataFrame(adjustments)
            if "登録日時" in df_adj.columns: df_adj = df_adj.sort_values("登録日時", ascending=False)
            ano_lot_map = {str(a.get("入荷No", "")).strip(): str(a.get("ロットNo", "")).strip() for a in arrivals}
            df_adj["ロットNo"] = df_adj.get("入荷No", "").astype(str).map(ano_lot_map)
            st.dataframe(fmt_df_numeric(df_adj[[c for c in ["登録日時", "調整日", "入荷No", "ロットNo", "調整袋数", "理由", "担当者"] if c in df_adj.columns]].head(200), ["調整袋数"]), use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════
#  🧹 資材管理
# ═══════════════════════════════════════════════════════════════
elif page == "🧹 資材管理":
    st.markdown('<div class="main-header"><h1>🧹 資材・消耗品管理</h1></div>', unsafe_allow_html=True)
    t_s1, t_s2 = st.tabs(["📋 在庫一覧・入出庫・棚卸", "🕒 ログ管理"])
    
    with t_s1:
        if not supplies: st.warning("資材が未登録です。マスタ設定よりご登録ください。")
        else:
            supply_inventory = get_supply_inventory()
            cat_list = sorted(list(set([str(s.get("カテゴリ") or "").strip() or "未分類" for s in supplies])))
            cat_options = ["📋 すべて表示"] + [f"{_deterministic_icon(c, _ICON_POOL)} {c}" for c in cat_list]

            st.markdown('<div style="font-weight:900; margin-bottom:8px;">① カテゴリを選択</div>', unsafe_allow_html=True)
            sel_cat_label = st.radio("カテゴリ", cat_options, horizontal=True, key="supply_cat_filter", label_visibility="collapsed")
            filtered_supplies = supplies if sel_cat_label == "📋 すべて表示" else [s for s in supplies if (str(s.get("カテゴリ") or "").strip() or "未分類") == cat_list[cat_options.index(sel_cat_label) - 1]]

            cols_grid = st.columns(min(3, len(filtered_supplies))) if filtered_supplies else []
            for idx, s in enumerate(filtered_supplies):
                sid = s.get("資材ID"); curr_qty = supply_inventory.get(sid, 0.0)
                with cols_grid[idx % 3]:
                    with st.container(border=True):
                        img_html = f'<img src="{s.get("画像URL")}" style="width:60px; height:60px; object-fit:cover; border-radius:8px; margin-right:12px; border:1px solid #e2e8f0;">' if s.get("画像URL") else '<div style="width:60px; height:60px; background:#e2e8f0; border-radius:8px; margin-right:12px; display:flex; align-items:center; justify-content:center; font-size:24px;">📦</div>'
                        st.markdown(f'<div style="display:flex; align-items:center; margin-bottom:12px;">{img_html}<div><div style="font-weight:900; font-size:1.1rem; color:#0f172a;">{s.get("資材名")}</div><div style="font-size:0.8rem; color:#64748b;">🏷️ {s.get("カテゴリ") or "未分類"}</div></div></div><div style="font-size:2.0rem; font-weight:900; color:#0f766e; margin-bottom:12px; text-align:center;">{fmt_kg(curr_qty)} <span style="font-size:1rem; color:#64748b;">個</span></div>', unsafe_allow_html=True)
                        
                        with st.popover("🔧 増減・棚卸調整", use_container_width=True):
                            adj_mode = st.radio("調整方法", ["➕➖ クイック入出庫", "📋 実地数量で確定"], key=f"sup_mode_{sid}", horizontal=True)
                            op_q = render_operator_selector(f"sup_qop_{sid}")
                            
                            def _sup_adj(s_id, delta, op_name, reason):
                                if hasattr(sheets, "append_supply_log"):
                                    sheets.append_supply_log({"ログID": f"LOG-{datetime.now().strftime('%Y%m%d%H%M%S%f')}", "登録日": str(date.today()), "資材ID": s_id, "処理": "入荷" if delta > 0 else "使用", "数量": abs(delta), "作業者": op_name, "備考": reason, "登録日時": datetime.now().isoformat()})
                            
                            if "クイック" in adj_mode:
                                qc1, qc2, qc3, qc4 = st.columns(4)
                                if qc1.button("➖10", key=f"sq_m10_{sid}", use_container_width=True): _sup_adj(sid, -10, op_q, "【クイック増減】"); st.toast("-10"); time.sleep(1); refresh()
                                if qc2.button("➖1", key=f"sq_m1_{sid}", use_container_width=True): _sup_adj(sid, -1, op_q, "【クイック増減】"); st.toast("-1"); time.sleep(1); refresh()
                                if qc3.button("➕1", key=f"sq_p1_{sid}", use_container_width=True): _sup_adj(sid, 1, op_q, "【クイック増減】"); st.toast("+1"); time.sleep(1); refresh()
                                if qc4.button("➕10", key=f"sq_p10_{sid}", use_container_width=True): _sup_adj(sid, 10, op_q, "【クイック増減】"); st.toast("+10"); time.sleep(1); refresh()
                            else:
                                actual_qty = st.number_input("実在庫数量", min_value=0.0, value=float(round(curr_qty, 2)), step=1.0, key=f"sup_actual_{sid}")
                                diff_qty = round(actual_qty - curr_qty, 4)
                                reason_txt = st.text_input("調整理由", key=f"sup_adj_reason_{sid}")
                                if st.button("💾 この実地数量で確定", type="primary", use_container_width=True, key=f"sup_adj_save_{sid}"):
                                    if diff_qty != 0:
                                        _sup_adj(sid, diff_qty, op_q, f"【棚卸調整:実地{fmt_kg(actual_qty)}に更新】{reason_txt}")
                                        st.success(f"現在庫を {fmt_kg(actual_qty)} に更新しました。"); time.sleep(1.5); refresh()

    with t_s2:
        if supply_logs:
            id_name_map = {s.get("資材ID"): s.get("資材名") for s in supplies}
            df_logs = pd.DataFrame(supply_logs)
            df_logs["資材名"] = df_logs["資材ID"].map(id_name_map)
            df_logs_sorted = df_logs.sort_values("登録日", ascending=False)
            st.dataframe(fmt_df_numeric(df_logs_sorted[["登録日", "資材名", "処理", "数量", "作業者", "備考"]].head(50), ["数量"]), use_container_width=True, hide_index=True)
            
            sec_title("🚨 ログの取り消し・削除")
            log_options = {f"{r.get('登録日','')} / {r.get('資材名','')} / {r.get('処理','')} {fmt_kg(r.get('数量',0))}": r.get("ログID", "") for _, r in df_logs_sorted.head(30).iterrows()}
            if log_options:
                sel_log = st.selectbox("削除するログを選択", list(log_options.keys()))
                if st.button("🗑️ このログを削除", type="primary"):
                    if hasattr(sheets, "delete_supply_log"):
                        sheets.delete_supply_log(log_options[sel_log])
                        st.success("削除しました。"); time.sleep(1); refresh()


# ═══════════════════════════════════════════════════════════════
#  🔍 トレース
# ═══════════════════════════════════════════════════════════════
elif page == "🔍 トレース":
    st.markdown('<div class="main-header"><h1>🔍 双方向原料トレース</h1></div>', unsafe_allow_html=True)
    trace_dir = st.radio("トレース方向", ["➡️ 原料ロットから製品を追跡（フォワード）", "⬅️ 製品から原料を遡及（バックワード）"])
    card_start()
    if "フォワード" in trace_dir:
        lot_list = sorted(list(set([str(a.get("ロットNo", "")).strip() for a in arrivals if a.get("ロットNo")])), reverse=True)
        tgt_lot = st.selectbox("検索する原料ロット", lot_list if lot_list else ["なし"])
        if st.button("➡️ 追跡開始", type="primary", use_container_width=True):
            match_brw = [b for b in brewing if any(tgt_lot in [re.sub(r'\(\d+%\)', '', x).strip() for x in str(item.get("lot", "")).split(",")] for item in parse_brewing_ingredients(b.get("その他添加物", "")))]
            if match_brw: st.dataframe(fmt_df_numeric(pd.DataFrame(match_brw)[["仕込日", "品名", "仕込量(kg)"]], ["仕込量(kg)"]), use_container_width=True, hide_index=True)
            else: st.warning("履歴がありません。")
    else:
        brw_opts = {f"No.{b.get('仕込No')} - {b.get('品名')} ({b.get('仕込日')})": b for b in brewing}
        if brw_opts:
            sel_b = st.selectbox("対象の製造記録", list(brw_opts.keys()))
            if st.button("⬅️ 遡及開始", type="primary", use_container_width=True):
                used_lots = [{"原料種別": ing.get("原料名"), "ロットNo": l} for ing in parse_brewing_ingredients(brw_opts[sel_b].get("その他添加物", "")) for l in [re.sub(r'\(\d+%\)', '', x).strip() for x in str(ing.get("lot", "")).split(",")] if l and l != "─"]
                if used_lots:
                    details = [{"原料種別": u["原料種別"], "ロットNo": u["ロットNo"], "入荷日": arr.get("入荷日") if (arr := next((a for a in arrivals if str(a.get("ロットNo", "")).strip() == u["ロットNo"]), None)) else "不明", "メーカー": arr.get("メーカー") if arr else "不明"} for u in used_lots]
                    st.dataframe(pd.DataFrame(details), use_container_width=True, hide_index=True)
                else: st.warning("原料ロットの記録はありません。")
    card_end()


# ═══════════════════════════════════════════════════════════════
#  📋 履歴・帳票
# ═══════════════════════════════════════════════════════════════
elif page == "📋 履歴・帳票":
    st.markdown('<div class="main-header"><h1>📋 製造履歴・帳票出力</h1></div>', unsafe_allow_html=True)
    if not brewing: st.info("データがありません。")
    else:
        df_brw = pd.DataFrame(brewing)
        df_brw["仕込日_dt"] = pd.to_datetime(df_brw["仕込日"], errors="coerce")
        card_start()
        c1, c2 = st.columns(2)
        s_date = c1.date_input("開始日", value=date.today().replace(day=1))
        filter_end_date = c2.date_input("終了日", value=date.today())
        
        filtered_df = df_brw[(df_brw["仕込日_dt"].dt.date >= s_date) & (df_brw["仕込日_dt"].dt.date <= filter_end_date)].copy().sort_values("仕込日", ascending=False)
        if HAS_OPENPYXL and not filtered_df.empty:
            wb = Workbook(); ws = wb.active; ws.title = "製造記録"
            for col_idx, h in enumerate(["製造日", "仕込No", "製品名", "担当者", "製造量(kg)", "石灰水(L)", "備考"], 1): ws.cell(row=1, column=col_idx, value=h)
            for r_idx, (_, row) in enumerate(filtered_df.iterrows(), 2):
                for c_idx, val in enumerate([row.get("仕込日", ""), row.get("仕込No", ""), row.get("品名", ""), row.get("メーカー", ""), float(row.get("仕込量(kg)", 0) or 0), float(row.get("石灰水(L)", 0) or 0), row.get("備考", "")], 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            excel_buffer = BytesIO(); wb.save(excel_buffer)
            st.download_button("🖨️ Excel帳票をダウンロード", data=excel_buffer.getvalue(), file_name=f"製造記録_{s_date}_{filter_end_date}.xlsx", type="primary")
        
        st.dataframe(fmt_df_numeric(filtered_df[["仕込日", "仕込No", "品名", "仕込量(kg)", "主原料ロット", "備考"]], ["仕込量(kg)"]), use_container_width=True, hide_index=True)
        card_end()
        
        card_start()
        sec_title("✏️ 製造履歴の一括編集（Excel風）")
        brw_editable_cols = [c for c in ["仕込日", "品名", "メーカー", "仕込量(kg)", "主原料ロット", "備考"] if c in filtered_df.columns]
        render_excel_history_editor(full_records=brewing, filtered_df=filtered_df.reset_index(drop=True), id_col="仕込No", editable_cols=brw_editable_cols, numeric_cols=["仕込量(kg)"], save_func=sheets.save_brewing, key_prefix="brw_hist")
        card_end()


# ═══════════════════════════════════════════════════════════════
#  📈 分析
# ═══════════════════════════════════════════════════════════════
elif page == "📈 分析":
    st.markdown('<div class="main-header"><h1>📈 製造・原料 分析</h1></div>', unsafe_allow_html=True)
    df_brw_global = pd.DataFrame(brewing)
    if df_brw_global.empty: st.info("データがありません。")
    else:
        df_brw_global["仕込日_dt"] = pd.to_datetime(df_brw_global["仕込日"], errors="coerce")
        df_brw_global["month"] = df_brw_global["仕込日_dt"].dt.to_period("M").astype(str)
        df_brw_global["仕込量(kg)"] = pd.to_numeric(df_brw_global["仕込量(kg)"], errors="coerce").fillna(0)
        
        card_start()
        monthly_trend = df_brw_global.groupby("month")["仕込量(kg)"].sum().reset_index().sort_values("month")
        fig = go.Figure(go.Bar(x=monthly_trend["month"], y=monthly_trend["仕込量(kg)"], marker_color="#0f766e"))
        fig.update_layout(title="月間生産推移 (kg)", xaxis_title="年月", yaxis_title="総製造量", plot_bgcolor="#ffffff")
        st.plotly_chart(fig, use_container_width=True)
        card_end()
        
        c1, c2 = st.columns(2)
        with c1:
            card_start()
            pie_data = df_brw_global.groupby("品名")["仕込量(kg)"].sum().reset_index().sort_values("仕込量(kg)", ascending=False)
            fig_tree = px.treemap(pie_data[pie_data["仕込量(kg)"] > 0], path=["品名"], values="仕込量(kg)", color="仕込量(kg)", color_continuous_scale=["#fde4d0", "#0f766e"], title="製品構成比")
            fig_tree.update_traces(texttemplate="<b>%{label}</b><br>%{value:,.0f} kg", textfont_size=14); fig_tree.update_layout(margin=dict(t=50, l=6, r=6, b=6))
            st.plotly_chart(fig_tree, use_container_width=True)
            card_end()
        with c2:
            card_start()
            topN = pie_data.sort_values("仕込量(kg)", ascending=True).tail(15)
            fig_bar = px.bar(topN, x="仕込量(kg)", y="品名", orientation='h', title="製造量 上位15品目", text="仕込量(kg)")
            fig_bar.update_traces(texttemplate="%{text:,.0f} kg", textposition="outside", marker_color="#0f766e"); fig_bar.update_layout(height=max(380, 34 * len(topN)), plot_bgcolor="#ffffff", yaxis_title="")
            st.plotly_chart(fig_bar, use_container_width=True)
            card_end()


# ═══════════════════════════════════════════════════════════════
#  ⚙️ マスタ設定
# ═══════════════════════════════════════════════════════════════
elif page == "⚙️ マスタ設定":
    st.markdown('<div class="main-header"><h1>⚙️ マスターデータ管理</h1></div>', unsafe_allow_html=True)
    t1, t2, t3, t4, t5 = st.tabs(["⚗️ 原料マスタ・詳細", "🏢 担当者", "🧪 レシピ・石灰", "📦 資材", "🏷️ グレード"])
    
    with t1:
        card_start()
        sec_title("🥦 原料名の登録・編集")
        ed_m = st.data_editor(pd.DataFrame({"原料名": pd.array([m for m in materials if not m.startswith("__")], dtype="string")}), num_rows="dynamic", use_container_width=True)
        if st.button("💾 原料マスタ保存", type="primary"):
            if hasattr(sheets, "save_materials"):
                sheets.save_materials([str(x).strip() for x in ed_m["原料名"].tolist() if x is not None and str(x).strip() and str(x).strip().lower() != "nan"])
                st.success("保存しました。"); time.sleep(1); refresh()
        card_end()

        if materials:
            card_start()
            sec_title("🔧 選択した原料の詳細設定")
            sel_m = st.selectbox("詳細を設定する原料", [m for m in materials if not m.startswith("__")])
            
            if sel_m:
                pt, wt = parse_op_data(order_points.get(sel_m, 0.0))
                active_lots = get_active_lots_from_master(order_points, sel_m)
                
                # 在庫があるロットのリスト
                hist_lots = [v["ロットNo"] for v in inventory_data.values() if v["原料種別"] == sel_m and v["現在庫(袋)"] > 0.01]
                # 万が一在庫ゼロでも既にピン留めされていたら消さないように追加
                for l in active_lots:
                    if l not in hist_lots: hist_lots.append(l)

                with st.form("mat_detail_form"):
                    c_pt, c_wt = st.columns(2)
                    new_pt = c_pt.number_input("🚨 発注点 (袋)", value=int(pt), step=1)
                    new_wt = c_wt.number_input("⚖️ 1袋重量 (kg)", value=int(wt), step=1)
                    
                    st.markdown("---")
                    st.markdown("##### 📦 製造で使用するロットの指定")
                    st.caption("仕込み画面で「メインの選択肢」として最前面に表示させたいロットのみを選んでください。指定しなかった在庫は「その他の在庫ありロット」に折りたたまれます。")
                    new_active_lots = st.multiselect("使用可能ロット", hist_lots, default=[x for x in active_lots if x in hist_lots])
                    
                    st.markdown("---")
                    st.markdown("##### 🖼️ 原料の画像")
                    uploaded_file = st.file_uploader("📷 画像 (任意)", type=["jpg", "png", "jpeg"])
                    
                    if st.form_submit_button("💾 詳細設定を保存する", type="primary"):
                        d = dict(order_points)
                        d[sel_m] = f"発注点:{new_pt}袋, 重量:{new_wt}kg"
                        d[f"__ACTIVE_LOTS_{sel_m}__"] = json.dumps(new_active_lots, ensure_ascii=False)
                        
                        if uploaded_file and HAS_PIL:
                            img = Image.open(uploaded_file); img.thumbnail((150, 150)); buffered = BytesIO(); img.save(buffered, format="PNG")
                            d[f"__IMAGE_{sel_m}__"] = f"data:image/png;base64,{base64.b64encode(buffered.getvalue()).decode('utf-8')}"
                        
                        if hasattr(sheets, "save_order_points"):
                            sheets.save_order_points(d)
                            st.success(f"{sel_m} の詳細設定を保存しました。"); time.sleep(1.5); refresh()
            card_end()

    with t2:
        card_start()
        ed_u = st.data_editor(pd.DataFrame({"担当者名": pd.array(inspectors, dtype="string")}), num_rows="dynamic", use_container_width=True)
        if st.button("💾 担当者保存", type="primary"):
            if hasattr(sheets, "save_inspectors"):
                sheets.save_inspectors([str(x).strip() for x in ed_u["担当者名"].tolist() if x is not None and str(x).strip() and str(x).strip().lower() != "nan"])
                st.success("保存しました。"); time.sleep(1); refresh()
        card_end()

    with t3:
        recipe_kind = st.radio("レシピ種別を選択", ["🍽️ 通常レシピ（仕込み配合）", "🌶️ 調味料レシピ（希釈）"], horizontal=True)
        if "通常" in recipe_kind:
            card_start()
            normal_recipes = [r for r in recipes_raw if r.get("大カテゴリ") != "調味料"]
            edit_mode = st.radio("操作", ["新規作成", "既存の編集"], horizontal=True)
            target_recipe, old_json = None, "[]"
            if edit_mode == "既存の編集" and normal_recipes:
                target_recipe = next((r for r in normal_recipes if r["品名"] == st.selectbox("編集するレシピ", [r["品名"] for r in normal_recipes])), None)
                if target_recipe: old_json = target_recipe.get("配合JSON", "[]")

            init_name = target_recipe["品名"] if target_recipe else ""
            init_cat_m = target_recipe.get("大カテゴリ") if target_recipe and target_recipe.get("大カテゴリ") in ["プラント", "OKM", "手詰め"] else "プラント"
            init_cat_s = target_recipe.get("中カテゴリ", "黒") if target_recipe else "黒"
            init_items = safe_parse_recipe(old_json)
            def_mats = ["(未設定)", "水"] + materials

            with st.form("recipe_form"):
                cat_main = st.radio("大カテゴリ", ["🏭 プラント", "🟦 OKM", "✋ 手詰め"], index=["プラント", "OKM", "手詰め"].index(init_cat_m), horizontal=True)
                cat_sub = st.radio("中カテゴリ", ["⚪ 白", "⚫ 黒", "❄️ 耐冷", "🍽️ ショクカイ", "🍜 めん", "📦 その他"], index=["白","黒","耐冷","ショクカイ","めん","その他"].index(init_cat_s) if init_cat_s in ["白","黒","耐冷","ショクカイ","めん","その他"] else 1, horizontal=True)
                new_p_name = st.text_input("製品名", value=init_name, disabled=(target_recipe is not None))
                
                cols_recipe = []
                for j in range(10):
                    c_n, c_w = st.columns([2, 1])
                    def_mat_val = init_items[j]["原料名"] if j < len(init_items) else "(未設定)"
                    def_rat_val = float(init_items[j]["比率"]) if j < len(init_items) else 0.00
                    ing_mat = c_n.selectbox(f"成分 {j+1}", def_mats, index=def_mats.index(def_mat_val) if def_mat_val in def_mats else 0, key=f"rmat_{j}")
                    ing_ratio = c_w.number_input("比率(％)", min_value=0.00, value=def_rat_val, step=0.01, key=f"rrat_{j}")
                    cols_recipe.append({"name": ing_mat, "ratio": ing_ratio})
                
                if st.form_submit_button("💾 レシピを保存"):
                    text_recipe = ", ".join([f"{i['name']}:{i['ratio']}%" for i in cols_recipe if i["name"] != "(未設定)" and i["ratio"] > 0])
                    cat_str = ["プラント", "OKM", "手詰め"][["🏭 プラント", "🟦 OKM", "✋ 手詰め"].index(cat_main)]
                    updated_recipes = [r for r in recipes_raw if r["品名"] != new_p_name]
                    updated_recipes.append({"品名": new_p_name, "大カテゴリ": cat_str, "中カテゴリ": cat_sub.split(" ")[1] if cat_str == "プラント" else "その他", "配合JSON": text_recipe})
                    if hasattr(sheets, "save_recipes"): sheets.save_recipes(updated_recipes); st.success("保存しました。"); time.sleep(1); refresh()
            card_end()
        else:
            card_start()
            sec_title("🌶️ 調味料 希釈レシピ設定")
            seasoning_recipes = [r for r in recipes_raw if r.get("大カテゴリ") == "調味料"]
            s_edit_mode = st.radio("操作", ["新規作成", "既存の編集"], horizontal=True)
            s_target, s_old_json = None, "[]"
            if s_edit_mode == "既存の編集" and seasoning_recipes:
                s_target = next((r for r in seasoning_recipes if r["品名"] == st.selectbox("編集する調味料", [r["品名"] for r in seasoning_recipes])), None)
                if s_target: s_old_json = s_target.get("配合JSON", "[]")

            s_init_name = s_target["品名"] if s_target else ""
            s_init_items = safe_parse_seasoning_recipe(s_old_json)
            season_def_mats = ["(未設定)"] + materials

            with st.form("seasoning_recipe_form"):
                s_new_name = st.text_input("調味料レシピ名", value=s_init_name, disabled=(s_target is not None))
                s_cols_recipe = []
                for j in range(5):
                    c_n, c_r = st.columns([2, 1])
                    def_mat_val = s_init_items[j]["原料名"] if j < len(s_init_items) else "(未設定)"
                    def_ratio_val = float(s_init_items[j]["希釈倍率"]) if j < len(s_init_items) else 1.0
                    ing_mat = c_n.selectbox(f"原料 {j+1}", season_def_mats, index=season_def_mats.index(def_mat_val) if def_mat_val in season_def_mats else 0, key=f"smat_{j}")
                    ing_dil = c_r.number_input("希釈倍率", min_value=0.1, value=def_ratio_val, step=0.1, key=f"sdil_{j}")
                    s_cols_recipe.append({"name": ing_mat, "dil": ing_dil})

                if st.form_submit_button("💾 調味料レシピを保存"):
                    if not s_new_name.strip(): st.error("レシピ名必須")
                    else:
                        text_s_recipe = ", ".join([f"{i['name']}:{i['dil']}倍" for i in s_cols_recipe if i["name"] != "(未設定)"])
                        updated_recipes = [r for r in recipes_raw if r["品名"] != s_new_name]
                        updated_recipes.append({"品名": s_new_name, "大カテゴリ": "調味料", "中カテゴリ": "希釈", "配合JSON": text_s_recipe})
                        if hasattr(sheets, "save_recipes"): sheets.save_recipes(updated_recipes); st.success("保存しました。"); time.sleep(1); refresh()
            card_end()
            
        card_start()
        sec_title("🌡️ 石灰の季節増量・調整ルール設定")
        lime_rows = [{"対象製品": "(全体デフォルト)", "開始月": int(parse_lime_config(order_points).get("start_month", 6)), "終了月": int(parse_lime_config(order_points).get("end_month", 9)), "増量割合(%)": float(parse_lime_config(order_points).get("add_ratio", 0.01)) * 100, "理由": parse_lime_config(order_points).get("reason", "夏場の高温対策")}]
        p_names = sorted([r.get("品名") for r in recipes_raw if r.get("大カテゴリ") != "調味料" and r.get("品名")])
        for p in p_names:
            if f"__LIME_CONFIG_{p}__" in order_points:
                cfg = parse_lime_config(order_points, product_name=p)
                lime_rows.append({"対象製品": p, "開始月": int(cfg.get("start_month", 6)), "終了月": int(cfg.get("end_month", 9)), "増量割合(%)": float(cfg.get("add_ratio", 0.01)) * 100, "理由": cfg.get("reason", "個別設定")})
        
        ed_lime = st.data_editor(pd.DataFrame(lime_rows), num_rows="dynamic", use_container_width=True, column_config={"対象製品": st.column_config.SelectboxColumn("対象製品", options=["(全体デフォルト)"] + p_names, required=True), "開始月": st.column_config.NumberColumn("開始月", min_value=1, max_value=12, step=1, required=True), "終了月": st.column_config.NumberColumn("終了月", min_value=1, max_value=12, step=1, required=True), "増量割合(%)": st.column_config.NumberColumn("増量割合(%)", min_value=0.0, step=0.1, required=True)})
        
        if st.button("💾 石灰ルールを保存", type="primary"):
            new_dict = {k: v for k, v in order_points.items() if not k.startswith("__LIME_CONFIG")}
            for _, r in ed_lime.iterrows():
                tgt = str(r.get("対象製品", "")).strip()
                if tgt and str(tgt).lower() != "nan":
                    new_dict["__LIME_CONFIG__" if tgt == "(全体デフォルト)" else f"__LIME_CONFIG_{tgt}__"] = f"開始:{int(r.get('開始月', 6))}月, 終了:{int(r.get('終了月', 9))}月, 割合:{float(r.get('増量割合(%)', 1.0)) / 100.0}, 理由:{str(r.get('理由', '夏場対策'))}"
            if hasattr(sheets, "save_order_points"): sheets.save_order_points(new_dict); st.success("更新しました。"); time.sleep(1.5); refresh()
        card_end()

    with t4:
        card_start()
        with st.form("new_sup_form"):
            new_s_name = st.text_input("資材名称 ＊")
            new_s_cat = st.text_input("カテゴリ (例: 包材)")
            uploaded_file = st.file_uploader("📷 画像 (任意)", type=["jpg", "png", "jpeg"])
            if st.form_submit_button("💾 資材を登録"):
                if not new_s_name: st.error("名称は必須です。")
                else:
                    img_str = ""
                    if uploaded_file and HAS_PIL:
                        img = Image.open(uploaded_file); img.thumbnail((150, 150)); buffered = BytesIO(); img.save(buffered, format="PNG")
                        img_str = f"data:image/png;base64,{base64.b64encode(buffered.getvalue()).decode('utf-8')}"
                    cur_sup = supplies.copy()
                    cur_sup.append({"資材ID": f"SUP-{datetime.now().strftime('%Y%m%d%H%M%S')}", "資材名": new_s_name, "カテゴリ": new_s_cat, "画像URL": img_str, "初期在庫": 0, "発注点": 10, "登録日": str(date.today())})
                    if hasattr(sheets, "save_supplies"): sheets.save_supplies(cur_sup); st.success("登録しました。"); time.sleep(1); refresh()
        card_end()

    with t5:
        card_start()
        sec_title("🏷️ こんにゃく粉 グレードマスタ")
        ed_grade = st.data_editor(pd.DataFrame({"グレード名": pd.array(parse_grade_list(order_points), dtype="string")}), num_rows="dynamic", use_container_width=True)
        if st.button("💾 グレードマスタ保存", type="primary"):
            save_grade_list(order_points, [str(x).strip() for x in ed_grade["グレード名"].tolist() if x is not None and str(x).strip() and str(x).strip().lower() != "nan"])
            st.success("保存しました。"); time.sleep(1); refresh()
        card_end()
