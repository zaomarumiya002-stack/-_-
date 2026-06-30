"""
report_generator.py  ─  Excel帳票出力モジュール（完全同期版）
"""
from pathlib import Path
from datetime import date, datetime
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("data/reports")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

C_HDR_BG  = "1e3a8a"
C_HDR_FG  = "ffffff"
C_TITLE   = "1e1b4b"
C_ROW_ALT = "f8fafc"
C_TOTAL   = "fef08a"
C_WARN    = "fee2e2"
C_BORDER  = "cbd5e1"

def _side(): return Side(style="thin", color=C_BORDER)
def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())

def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Segoe UI", bold=True, color=C_HDR_FG, size=10)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, val, align="left", bg=None, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Segoe UI", size=9, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border    = _border()
    if bg:      c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt: c.number_format = num_fmt
    return c

def _title(ws, title, subtitle=""):
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 16
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name="Segoe UI", bold=True, size=14, color=C_TITLE)
    if subtitle:
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = Font(name="Segoe UI", size=9, color="475569")
    d = ws.cell(row=2, column=8, value=f"出力日: {date.today().strftime('%Y/%m/%d')}")
    d.font = Font(name="Segoe UI", size=9, color="475569")
    d.alignment = Alignment(horizontal="right")

def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

# ═══════════════════════════════════════════════════════════════
# 入荷記録帳票
# ═══════════════════════════════════════════════════════════════
def generate_arrival_report(arrivals: list) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入荷記録"
    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 2

    _title(ws, "📦 原料入荷記録一覧", f"総件数: {len(arrivals)} 件")

    HDRS = [
        ("入荷No",12),("入荷日",12),("メーカー",16),("ロットNo",16),
        ("原料種別",16),("袋数",8),("1袋重量(kg)",12),("総量(kg)",12),
        ("外観",10),("品名・規格確認",14),("賞味期限",12),("異物",10),
        ("異常内容",20),("担当者",10),("備考",20)
    ]
    HR = 4
    ws.row_dimensions[HR].height = 26
    for ci, (h, w) in enumerate(HDRS, start=2):
        _hdr(ws, HR, ci, h, w)

    for ri, a in enumerate(reversed(arrivals), start=HR + 1):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 20
        
        row = [
            (a.get("入荷No",""), "center", None),
            (a.get("入荷日",""), "center", None),
            (a.get("メーカー",""), "left", None),
            (a.get("ロットNo",""), "center", None),
            (a.get("原料種別",""), "left", None),
            (float(a.get("袋数", 0)), "right", "#,##0"),
            (float(a.get("1袋重量(kg)", 20)), "right", "#,##0.0"),
            (float(a.get("総量(kg)", 0)), "right", "#,##0.0"),
            (a.get("外観",""), "center", None),
            (a.get("品名・規格確認",""), "center", None),
            (a.get("賞味期限",""), "center", None),
            (a.get("異物",""), "center", None),
            (a.get("異常内容","") or "─", "left", None),
            (a.get("担当者",""), "center", None),
            (a.get("備考","") or "─", "left", None),
        ]
        
        for ci, (val, align, nf) in enumerate(row, start=2):
            row_bg = C_WARN if "NG" in str(val) else bg
            _cell(ws, ri, ci, val, align, bg=row_bg, num_fmt=nf)

    ws.freeze_panes = f"B{HR+1}"
    path = OUTPUT_DIR / f"入荷記録_{_ts()}.xlsx"
    wb.save(path)
    return path

# ═══════════════════════════════════════════════════════════════
# 仕込み記録帳票
# ═══════════════════════════════════════════════════════════════
def generate_brewing_report(brewing: list) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "仕込み記録"
    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 2

    _title(ws, "🧪 仕込み・製造実績記録", f"総件数: {len(brewing)} 件")

    HDRS = [
        ("仕込No",10),("仕込日",12),("品名",20),("メーカー",14),("主原料ロット",16),
        ("仕込量(kg)",12),("こんにゃく精粉(kg)",16),("海藻粉(kg)",12),("海藻粉ロット",16),
        ("デンプン(kg)",12),("デンプンロット",16),("石灰(kg)",10),("備考",20)
    ]
    HR = 4
    ws.row_dimensions[HR].height = 26
    for ci, (h, w) in enumerate(HDRS, start=2):
        _hdr(ws, HR, ci, h, w)

    tot_brew = tot_konjac = tot_seaweed = 0.0
    for ri, b in enumerate(reversed(brewing), start=HR + 1):
        bg = C_ROW_ALT if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 20
        
        b_val = float(b.get("仕込量(kg)", 0) or 0)
        k_val = float(b.get("こんにゃく精粉(kg)", 0) or 0)
        s_val = float(b.get("海藻粉(kg)", 0) or 0)
        tot_brew += b_val
        tot_konjac += k_val
        tot_seaweed += s_val

        row = [
            (b.get("仕込No",""), "center", None),
            (b.get("仕込日",""), "center", None),
            (b.get("品名",""), "left", None),
            (b.get("メーカー",""), "left", None),
            (b.get("主原料ロット",""), "center", None),
            (b_val, "right", "#,##0.0"),
            (k_val, "right", "#,##0.0"),
            (s_val, "right", "#,##0.0"),
            (b.get("海藻粉ロット","") or "─", "center", None),
            (float(b.get("デンプン(kg)", 0) or 0), "right", "#,##0.0"),
            (b.get("デンプンロット","") or "─", "center", None),
            (float(b.get("石灰(kg)", 0) or 0), "right", "#,##0.00"),
            (b.get("備考","") or "─", "left", None),
        ]
        for ci, (val, align, nf) in enumerate(row, start=2):
            _cell(ws, ri, ci, val, align, bg=bg, num_fmt=nf)

    # 合計行
    tr = HR + len(brewing) + 1
    ws.row_dimensions[tr].height = 22
    _cell(ws, tr, 2, "総合計", "center", bg=C_TOTAL, bold=True)
    _cell(ws, tr, 7, tot_brew, "right", bg=C_TOTAL, num_fmt="#,##0.0", bold=True)
    _cell(ws, tr, 8, tot_konjac, "right", bg=C_TOTAL, num_fmt="#,##0.0", bold=True)
    _cell(ws, tr, 9, tot_seaweed, "right", bg=C_TOTAL, num_fmt="#,##0.0", bold=True)

    ws.freeze_panes = f"B{HR+1}"
    path = OUTPUT_DIR / f"仕込み記録_{_ts()}.xlsx"
    wb.save(path)
    return path
